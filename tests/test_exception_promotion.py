from __future__ import annotations

import sys
from datetime import date

from openpyxl import Workbook, load_workbook

import promote_reviewed_exceptions as promote_cli
from src.exception_promotion import (
    EXCEPTION_SHEET_NAME,
    EXCEPTION_STATUS_COLUMN,
    FLOW_COLUMN,
    MISSING_ERROR_COLUMN,
    STATUS_INVALID_FLOW,
    STATUS_MISSING_DATA,
    STATUS_PROMOTED,
    promote_reviewed_exceptions,
)
from src.flows import FLOW_BAO_NO, FLOW_CHI_TIEN_MAT, FLOW_THU_TIEN_MAT
from src.models import ProcessedTransaction
from src.output_writer import (
    EXCEPTION_COLUMNS,
    EXCEPTION_TECHNICAL_COLUMNS,
    RPA_BUSINESS_COLUMNS,
    RPA_CHI_TIEN_MAT_COLUMNS,
    RPA_REASON_UNICODE_COLUMN,
    RPA_THU_TIEN_MAT_COLUMNS,
    write_excel,
)
from src.rpa_input_status import INPUT_MESSAGE_COLUMN, INPUT_STATUS_COLUMN, INPUT_UPDATED_AT_COLUMN, RPA_INPUT_SHEETS
from src.rpa_tracking import STATUS_PENDING


def _input_columns(sheet_name: str, *, legacy_cash_headers: bool = False) -> list[str]:
    if sheet_name == "THU_TIEN_MAT_INPUT":
        columns = list(RPA_THU_TIEN_MAT_COLUMNS)
        if legacy_cash_headers:
            columns[columns.index("Người nộp tiền")] = "Người nhận tiền"
    elif sheet_name == "CHI_TIEN_MAT_INPUT":
        columns = list(RPA_CHI_TIEN_MAT_COLUMNS)
        if legacy_cash_headers:
            columns[columns.index("Người nhận tiền")] = "Người nộp tiền"
    else:
        columns = list(RPA_BUSINESS_COLUMNS)
    if RPA_REASON_UNICODE_COLUMN not in columns:
        columns.insert(columns.index("Lí do") + 1, RPA_REASON_UNICODE_COLUMN)
    return columns


def _make_workbook(path, *, legacy: bool = False) -> None:
    workbook = Workbook()
    for index, sheet_name in enumerate(RPA_INPUT_SHEETS):
        ws = workbook.active if index == 0 else workbook.create_sheet(sheet_name)
        ws.title = sheet_name
        ws.append(_input_columns(sheet_name, legacy_cash_headers=legacy))
    exception_ws = workbook.create_sheet(EXCEPTION_SHEET_NAME)
    if legacy:
        exception_ws.append(
            [
                "Mã định danh", "File gốc", "Sheet gốc", "Dòng gốc", "Ngân hàng", "Luồng", "Ngày CT",
                "Nội dung giao dịch gốc", "Người hưởng/Người chuyển", "Người nhận tiền", "Người nộp tiền",
                "Mã ĐT", "Tên ĐT suy luận", "Lí do", RPA_REASON_UNICODE_COLUMN, "TK nợ", "TK có",
                "Thành tiền", "Ngoại tệ", "Số tiền ngoại tệ", "Tỷ giá", "transaction_uid", "run_id",
                "Duyệt nhập RPA", FLOW_COLUMN, "Trạng thái xử lý exception", "Lỗi còn thiếu",
                "Promoted to input", "Promoted sheet", "Promoted at",
            ]
        )
    else:
        exception_ws.append(EXCEPTION_COLUMNS)
    workbook.save(path)


def _append_exception(path, *, legacy: bool = False, **overrides) -> None:
    workbook = load_workbook(path)
    ws = workbook[EXCEPTION_SHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    if legacy:
        row = {
            "Mã định danh": "uid_exception", "File gốc": "statement.xlsx", "Sheet gốc": "Sheet1",
            "Dòng gốc": 2, "Ngân hàng": "ACB", "Luồng": FLOW_BAO_NO, FLOW_COLUMN: FLOW_BAO_NO,
            "Ngày CT": date(2026, 6, 20), "Nội dung giao dịch gốc": "TT ABC",
            "Người hưởng/Người chuyển": "ABC", "Người nhận tiền": "Legacy person",
            "Người nộp tiền": "Legacy person", "Mã ĐT": "ABC", "Tên ĐT suy luận": "ABC",
            RPA_REASON_UNICODE_COLUMN: "Thanh toán ABC", "TK nợ": "331", "TK có": "1121CT",
            "Thành tiền": 1000, "transaction_uid": "uid_exception", "run_id": "run1",
            "Duyệt nhập RPA": "yes",
        }
    else:
        row = {
            "Duyệt nhập RPA": "yes", EXCEPTION_STATUS_COLUMN: "Chưa duyệt", MISSING_ERROR_COLUMN: "",
            "Nguồn sao kê": "statement.xlsx | Sheet1 | dòng 2", "Ngân hàng": "ACB", FLOW_COLUMN: FLOW_BAO_NO,
            "Ngày CT": date(2026, 6, 20), "Nội dung giao dịch": "TT ABC", "Đối tượng giao dịch": "ABC",
            "Người nộp/nhận tiền": "Nguyen Van A", "Mã ĐT": "ABC", "Tên ĐT": "ABC",
            "Lý do": "Thanh toán ABC", "TK nợ": "331", "TK có": "1121CT", "Thành tiền": 1000,
            "transaction_uid": "uid_exception", "run_id": "run1", "source_file": "statement.xlsx",
            "source_sheet": "Sheet1", "source_row": 2,
        }
    row.update(overrides)
    ws.append([row.get(header, "") for header in headers])
    workbook.save(path)


def _rows(path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    ws = workbook[sheet_name]
    headers = [cell.value for cell in ws[1]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]


def test_generated_exception_sheet_is_simplified_and_hides_technical_columns(tmp_path):
    output_file = tmp_path / "rpa_input.xlsx"
    item = ProcessedTransaction(
        source_file="statement.xlsx", original_row_index=2, bank="ACB", flow=FLOW_BAO_NO,
        transaction_date=date(2026, 6, 20), object_code="ERROR", object_name="", reason="Cần bổ sung",
        debit_account="", credit_account="1121CT", amount=1000, use_case="", original_content="TT ABC",
        counterparty_raw="ABC", doc_no="REF1", status="ERROR", error_note="Thiếu mã đối tượng",
        confidence=0, transaction_uid="uid_error", source_sheet="Sheet1",
    )
    write_excel([item], output_file, run_id="run1", rpa_reason_encoding="tcvn3")

    workbook = load_workbook(output_file)
    ws = workbook[EXCEPTION_SHEET_NAME]
    assert [cell.value for cell in ws[1]] == EXCEPTION_COLUMNS
    headers = {cell.value: cell.column for cell in ws[1]}
    for name in EXCEPTION_TECHNICAL_COLUMNS:
        assert ws.column_dimensions[ws.cell(1, headers[name]).column_letter].hidden
    assert len(ws.data_validations.dataValidation) == 2


def test_promotes_new_exception_and_is_idempotent(tmp_path):
    input_file = tmp_path / "rpa_input.xlsx"
    _make_workbook(input_file)
    _append_exception(input_file)

    first = promote_reviewed_exceptions(input_file)
    second = promote_reviewed_exceptions(input_file)

    assert first.promoted == 1
    assert second.promoted == 0
    assert second.skipped_already_promoted == 1
    row = _rows(input_file, "BAO_NO_INPUT")[0]
    assert row["Mã ĐT"] == "ABC"
    assert row["transaction_uid"] == "uid_exception"
    assert row[INPUT_STATUS_COLUMN] == STATUS_PENDING
    assert row[INPUT_MESSAGE_COLUMN] in ("", None)
    assert row[INPUT_UPDATED_AT_COLUMN] in ("", None)
    exception = _rows(input_file, EXCEPTION_SHEET_NAME)[0]
    assert str(exception[EXCEPTION_STATUS_COLUMN]).startswith(f"{STATUS_PROMOTED} vào BAO_NO_INPUT")
    assert exception[MISSING_ERROR_COLUMN] in ("", None)


def test_missing_required_data_and_invalid_flow_are_actionable(tmp_path, monkeypatch, capsys):
    input_file = tmp_path / "rpa_input.xlsx"
    _make_workbook(input_file)
    _append_exception(input_file, **{"Mã ĐT": "", "transaction_uid": "uid_missing"})
    monkeypatch.setattr(sys, "argv", ["promote_reviewed_exceptions.py", "--input-file", str(input_file), "--fail-on-validation-error"])
    assert promote_cli.main() == 2
    assert "Failed validation 1 rows." in capsys.readouterr().out
    exception = _rows(input_file, EXCEPTION_SHEET_NAME)[0]
    assert exception[EXCEPTION_STATUS_COLUMN] == STATUS_MISSING_DATA
    assert "Thiếu Mã ĐT" in exception[MISSING_ERROR_COLUMN]

    invalid_file = tmp_path / "invalid.xlsx"
    _make_workbook(invalid_file)
    _append_exception(invalid_file, **{FLOW_COLUMN: "khong_hop_le"})
    result = promote_reviewed_exceptions(invalid_file)
    assert result.failed_validation == 1
    invalid = _rows(invalid_file, EXCEPTION_SHEET_NAME)[0]
    assert invalid[EXCEPTION_STATUS_COLUMN] == STATUS_INVALID_FLOW


def test_cash_promotions_use_correct_person_columns(tmp_path):
    input_file = tmp_path / "rpa_input.xlsx"
    _make_workbook(input_file)
    _append_exception(
        input_file, **{"transaction_uid": "uid_thu", FLOW_COLUMN: FLOW_THU_TIEN_MAT,
                       "TK nợ": "1111", "TK có": "1121CT", "Người nộp/nhận tiền": "Tran Thi Thu"}
    )
    _append_exception(
        input_file, **{"transaction_uid": "uid_chi", FLOW_COLUMN: FLOW_CHI_TIEN_MAT,
                       "TK nợ": "1121CT", "TK có": "1111", "Người nộp/nhận tiền": "Tran Van Chi"}
    )
    result = promote_reviewed_exceptions(input_file)
    assert result.promoted == 2
    assert _rows(input_file, "THU_TIEN_MAT_INPUT")[0]["Người nộp tiền"] == "Tran Thi Thu"
    assert _rows(input_file, "CHI_TIEN_MAT_INPUT")[0]["Người nhận tiền"] == "Tran Van Chi"


def test_legacy_exception_without_rpa_tasks_still_promotes_and_normalizes_cash_header(tmp_path):
    input_file = tmp_path / "legacy.xlsx"
    _make_workbook(input_file, legacy=True)
    _append_exception(
        input_file, legacy=True, **{"transaction_uid": "uid_legacy", FLOW_COLUMN: FLOW_THU_TIEN_MAT,
                                   "TK nợ": "1111", "TK có": "1121CT", "Người nhận tiền": "Legacy Thu"}
    )
    result = promote_reviewed_exceptions(input_file)
    assert result.promoted == 1
    workbook = load_workbook(input_file, data_only=True)
    assert "RPA_TASKS" not in workbook.sheetnames
    assert "Người nộp tiền" in [cell.value for cell in workbook["THU_TIEN_MAT_INPUT"][1]]
    assert _rows(input_file, "THU_TIEN_MAT_INPUT")[0]["Người nộp tiền"] == "Legacy Thu"
