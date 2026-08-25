from __future__ import annotations

import json
import sys
from datetime import date

import pandas as pd
from openpyxl import load_workbook

import update_rpa_status as update_cli
from src.models import ProcessedTransaction
from src.output_writer import write_outputs
from src.rpa_summary import (
    COMPLETED_COLUMN,
    MESSAGE_COLUMN,
    STATUS_COLUMN,
    SUMMARY_COLUMNS,
    SUMMARY_SHEET_NAME,
    SUMMARY_TECHNICAL_COLUMNS,
    VOUCHER_COLUMN,
    abort_rpa_run,
    mark_rpa_done,
    mark_rpa_error,
    mark_rpa_started,
    reset_all_rpa_status,
    write_summary,
)
from src.rpa_tracking import STATUS_DONE, STATUS_PENDING, reset_all_records


def _processed(uid: str, row_index: int, amount: int = 1000) -> ProcessedTransaction:
    return ProcessedTransaction(
        source_file="sample.xlsx", original_row_index=row_index, bank="ACB", flow="bao_no",
        transaction_date=date(2026, 4, 1), object_code="ABC", object_name="ABC",
        reason="Thanh toán ABC", debit_account="331", credit_account="1121CT", amount=amount,
        use_case="Chi phí thanh toán", original_content=f"TT CHO ABC {row_index}", counterparty_raw="ABC",
        doc_no=str(row_index), status="OK", error_note="", confidence=0.95,
        transaction_uid=uid, source_sheet="Statement",
    )


def _new_summary_row(uid: str, status: str = STATUS_PENDING) -> dict[str, object]:
    row = {column: "" for column in SUMMARY_COLUMNS}
    row.update(
        {
            "Ngày CT": date(2026, 4, 1), "Ngân hàng": "ACB", "Luồng": "bao_no",
            "Nội dung giao dịch": "TT CHO ABC", "Mã ĐT": "ABC", "Tên ĐT": "ABC",
            "TK nợ": "331", "TK có": "1121CT", "Thành tiền": 1000,
            "Kết quả phân loại": "OK", STATUS_COLUMN: status,
            "Nguồn sao kê": "sample.xlsx | Statement | dòng 2", "transaction_uid": uid,
        }
    )
    return row


def _legacy_summary(path, rows: list[dict[str, object]]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(rows).to_excel(writer, sheet_name=SUMMARY_SHEET_NAME, index=False)
        pd.DataFrame([{"rpa_status": "chua_nhap", "count": len(rows)}]).to_excel(
            writer, sheet_name="STATUS_COUNTS", index=False
        )


def _read(path) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=SUMMARY_SHEET_NAME, dtype=object).fillna("")


def test_legacy_summary_is_backed_up_migrated_and_completed_rows_are_filtered(tmp_path):
    summary_path = tmp_path / "rpa_summary.xlsx"
    _legacy_summary(
        summary_path,
        [
            {
                "transaction_uid": "uid_done", "source_file": "sample.xlsx", "source_sheet": "Statement",
                "source_row_index": 2, "bank": "ACB", "flow": "bao_no", "transaction_date": date(2026, 4, 1),
                "original_content": "TT CHO ABC 2", "amount": 1000, "object_code": "ABC", "object_name": "ABC",
                "debit_account": "331", "credit_account": "1121CT", "status": STATUS_DONE,
                "voucher_no": "BN001", "completed_at": "2026-04-01T08:01:00", "last_run_id": "old_run",
            }
        ],
    )
    result = write_outputs([_processed("uid_done", 2), _processed("uid_new", 3)], tmp_path, {"output": {}})

    input_wb = load_workbook(result.excel_path, data_only=True)
    assert input_wb.sheetnames == ["BAO_NO_INPUT", "BAO_CO_INPUT", "THU_TIEN_MAT_INPUT", "CHI_TIEN_MAT_INPUT", "EXCEPTION"]
    assert input_wb["BAO_NO_INPUT"].max_row == 2
    summary_wb = load_workbook(summary_path)
    assert summary_wb.sheetnames == [SUMMARY_SHEET_NAME]
    rows = _read(summary_path).set_index("transaction_uid")
    assert rows.at["uid_done", STATUS_COLUMN] == STATUS_DONE
    assert rows.at["uid_done", VOUCHER_COLUMN] == "BN001"
    assert rows.at["uid_new", STATUS_COLUMN] == STATUS_PENDING
    backups = list((tmp_path / "backup").glob("rpa_summary_before_simplify_*.xlsx"))
    assert len(backups) == 1

    write_outputs([_processed("uid_new", 3)], tmp_path, {"output": {}})
    assert len(list((tmp_path / "backup").glob("rpa_summary_before_simplify_*.xlsx"))) == 1


def test_summary_ignores_legacy_tracking_json_and_keeps_history(tmp_path):
    tracking_path = tmp_path / "rpa_tracking.json"
    tracking_record = {"transaction_uid": "uid_excel_wins", "rpa_status": STATUS_DONE}
    tracking_path.write_text(json.dumps([tracking_record]), encoding="utf-8")
    result = write_outputs([_processed("uid_excel_wins", 2)], tmp_path, {"output": {}})
    rows = _read(result.summary_path).set_index("transaction_uid")
    assert rows.at["uid_excel_wins", STATUS_COLUMN] == STATUS_PENDING
    assert json.loads(tracking_path.read_text(encoding="utf-8")) == [tracking_record]


def test_summary_status_helpers_and_cli_keep_public_interface(tmp_path, monkeypatch):
    summary_path = tmp_path / "rpa_summary.xlsx"
    write_summary(pd.DataFrame([_new_summary_row("uid_pending")], columns=SUMMARY_COLUMNS), summary_path)

    mark_rpa_started(summary_path, "uid_pending", "run_1")
    started = _read(summary_path).set_index("transaction_uid").loc["uid_pending"]
    assert started[STATUS_COLUMN] == STATUS_PENDING
    assert started["rpa_started_at"]

    mark_rpa_done(summary_path, "uid_pending", "run_1", voucher_no="BN001", message="Đã nhập")
    done = _read(summary_path).set_index("transaction_uid").loc["uid_pending"]
    assert done[STATUS_COLUMN] == STATUS_DONE
    assert done[VOUCHER_COLUMN] == "BN001"
    assert done[MESSAGE_COLUMN] == "Đã nhập"
    assert done[COMPLETED_COLUMN]

    monkeypatch.setattr(
        sys,
        "argv",
        ["update_rpa_status.py", "--output-dir", str(tmp_path), "--uid", "uid_pending", "--status", STATUS_PENDING],
    )
    assert update_cli.main() == 0
    assert _read(summary_path).set_index("transaction_uid").at["uid_pending", STATUS_COLUMN] == STATUS_PENDING


def test_abort_and_reset_preserve_expected_workflow(tmp_path):
    summary_path = tmp_path / "rpa_summary.xlsx"
    rows = [_new_summary_row("uid_attempted"), _new_summary_row("uid_untouched"), _new_summary_row("uid_done", STATUS_DONE)]
    rows[2][VOUCHER_COLUMN] = "OLD"
    rows[2][COMPLETED_COLUMN] = "2026-04-01T08:01:00"
    write_summary(pd.DataFrame(rows, columns=SUMMARY_COLUMNS), summary_path)

    mark_rpa_started(summary_path, "uid_attempted", "run_abort")
    mark_rpa_error(summary_path, "uid_attempted", "run_abort", "VACOM row error")
    abort_rpa_run(summary_path, "run_abort", message="PAD abort")
    result = _read(summary_path).set_index("transaction_uid")
    assert result.at["uid_attempted", STATUS_COLUMN] == STATUS_PENDING
    assert result.at["uid_attempted", MESSAGE_COLUMN] == "PAD abort"
    assert result.at["uid_untouched", MESSAGE_COLUMN] == ""
    assert result.at["uid_done", STATUS_COLUMN] == STATUS_DONE
    assert result.at["uid_done", VOUCHER_COLUMN] == "OLD"

    reset_all_rpa_status(summary_path, message="Reset all")
    reset = _read(summary_path)
    assert set(reset[STATUS_COLUMN]) == {STATUS_PENDING}
    assert set(reset[MESSAGE_COLUMN]) == {"Reset all"}
    assert set(reset[VOUCHER_COLUMN]) == {""}


def test_run_level_cli_commands_keep_their_existing_flags(tmp_path, monkeypatch):
    summary_path = tmp_path / "rpa_summary.xlsx"
    write_summary(pd.DataFrame([_new_summary_row("uid_run")], columns=SUMMARY_COLUMNS), summary_path)

    mark_rpa_started(summary_path, "uid_run", "run_cli")
    monkeypatch.setattr(
        sys,
        "argv",
        ["update_rpa_status.py", "--output-dir", str(tmp_path), "--abort-run", "--run-id", "run_cli"],
    )
    assert update_cli.main() == 0

    monkeypatch.setattr(
        sys,
        "argv",
        ["update_rpa_status.py", "--output-dir", str(tmp_path), "--finalize-run", "--run-id", "run_cli"],
    )
    assert update_cli.main() == 0

    monkeypatch.setattr(
        sys,
        "argv",
        ["update_rpa_status.py", "--output-dir", str(tmp_path), "--reset-all", "--message", "Reset CLI"],
    )
    assert update_cli.main() == 0
    row = _read(summary_path).set_index("transaction_uid").loc["uid_run"]
    assert row[STATUS_COLUMN] == STATUS_PENDING
    assert row[MESSAGE_COLUMN] == "Reset CLI"


def test_summary_workbook_is_one_sheet_and_hides_technical_columns(tmp_path):
    path = tmp_path / "rpa_summary.xlsx"
    write_summary(pd.DataFrame([_new_summary_row("uid_pending")], columns=SUMMARY_COLUMNS), path)
    workbook = load_workbook(path)
    assert workbook.sheetnames == [SUMMARY_SHEET_NAME]
    ws = workbook[SUMMARY_SHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    assert headers == SUMMARY_COLUMNS
    assert ws.freeze_panes == "A2"
    header_map = {cell.value: cell.column_letter for cell in ws[1]}
    for column in SUMMARY_TECHNICAL_COLUMNS:
        assert ws.column_dimensions[header_map[column]].hidden
    assert ws[1][headers.index(STATUS_COLUMN)].comment is not None
    assert any("chua_nhap,hoan_thanh" in validation.formula1 for validation in ws.data_validations.dataValidation)


def test_reset_all_records_low_level_contract_is_unchanged():
    records = [
        {"transaction_uid": "done", "rpa_status": STATUS_DONE, "voucher_no": "BN1", "completed_at": "x"},
        {"transaction_uid": "pending", "rpa_status": STATUS_PENDING, "last_attempt_result": "error"},
    ]
    reset = reset_all_records(records, message="Reset all")
    assert {record["rpa_status"] for record in reset} == {STATUS_PENDING}
    assert {record["rpa_message"] for record in reset} == {"Reset all"}
    assert {record["voucher_no"] for record in reset} == {""}
