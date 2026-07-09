from __future__ import annotations

from datetime import date
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from src.config_loader import load_config, load_rules
from src.entity_extractor import EntityExtractor, OwnCompanyConfig
from src.flows import FLOW_BAO_CO, FLOW_BAO_NO, FLOW_CHI_TIEN_MAT, FLOW_THU_TIEN_MAT
from src.models import Transaction
from src.object_aliases import load_object_aliases
from src.object_matcher import ObjectMatcher
from src.object_overrides import load_object_overrides
from src.output_writer import (
    RPA_BUSINESS_COLUMNS,
    RPA_CHI_TIEN_MAT_COLUMNS,
    RPA_REASON_UNICODE_COLUMN,
    RPA_THU_TIEN_MAT_COLUMNS,
    write_excel,
    write_outputs,
)
from src.parsers.acb_parser import ACBParser
from src.parsers.msb_parser import MSBParser
from src.parsers.vcb_parser import VCBParser
from src.processor import process_all, process_transaction
from src.rule_engine import RuleEngine


PROJECT_ROOT = Path(__file__).resolve().parents[1]
STATEMENTS_DIR = PROJECT_ROOT / "input" / "statements"
INTERNAL_RECORDS = [
    {"code": "DUC", "name": "Lê Ngọc Đức"},
    {"code": "HOANGANH", "name": "Trịnh Hoàng Anh"},
    {"code": "HOA", "name": "Lê Thị Thanh Hoa"},
    {"code": "VIETHUNG", "name": "Phạm Việt Hùng"},
]
INTERNAL_ALIASES = {
    "DUC": ["LE NGOC DUC", "NGOC DUC"],
    "HOANGANH": ["TRINH HOANG ANH", "HOANG ANH"],
    "HOA": ["LE THI THANH HOA", "THANH HOA"],
    "VIETHUNG": ["PHAM VIET HUNG", "VIET HUNG"],
}


def _statement_sample(filename: str) -> Path:
    for base in [STATEMENTS_DIR, PROJECT_ROOT / "input"]:
        candidate = base / filename
        if candidate.exists():
            return candidate
    return STATEMENTS_DIR / filename


def _config() -> dict:
    config = load_config(PROJECT_ROOT / "config" / "config.yaml")
    config["ml"]["enabled"] = False
    return config


def _engine() -> RuleEngine:
    rules, _ = load_rules(None, PROJECT_ROOT / "config" / "default_rules.yaml")
    return RuleEngine(rules)


def _txn(description: str, debit: float = 0, credit: float = 0, bank: str = "VCB", counterparty: str = "") -> Transaction:
    return Transaction(
        source_file="sample.xlsx",
        bank=bank,
        transaction_date=date(2026, 4, 1),
        doc_no="REF1",
        description=description,
        counterparty_raw=counterparty,
        debit_amount=debit,
        credit_amount=credit,
        original_row_index=2,
    )


def _process(
    description: str,
    debit: float = 0,
    credit: float = 0,
    bank: str = "VCB",
    counterparty: str = "",
    receivable: list[dict] | None = None,
    payable: list[dict] | None = None,
    internal: list[dict] | None = None,
):
    return process_transaction(
        _txn(description, debit=debit, credit=credit, bank=bank, counterparty=counterparty),
        _config(),
        _engine(),
        ObjectMatcher.from_records(receivable or []),
        ObjectMatcher.from_records(payable or []),
        internal_matcher=ObjectMatcher.from_records(internal or INTERNAL_RECORDS, aliases=INTERNAL_ALIASES),
    )


@lru_cache(maxsize=1)
def _real_processing_context():
    own_company = OwnCompanyConfig.from_yaml(PROJECT_ROOT / "config" / "own_company.yaml")
    aliases = load_object_aliases(PROJECT_ROOT / "config" / "object_aliases.yaml")
    overrides = load_object_overrides(PROJECT_ROOT / "config" / "object_overrides.yaml")

    def merge_aliases(catalog: str) -> dict[str, list[str]]:
        merged: dict[str, list[str]] = {}
        for section in [aliases.get(catalog, {}), overrides.get(catalog, {}).get("aliases", {})]:
            for code, values in section.items():
                merged.setdefault(code, []).extend(values or [])
        return {code: list(dict.fromkeys(values)) for code, values in merged.items()}

    receivable = ObjectMatcher.from_excel(
        PROJECT_ROOT / "input" / "R_DMDT1 1.xlsx",
        aliases=merge_aliases("receivable"),
        exact_phrase_overrides=overrides.get("receivable", {}).get("exact_phrases", {}),
        supplemental_objects=overrides.get("receivable", {}).get("supplemental_objects", []),
        own_company=own_company,
    )
    payable = ObjectMatcher.from_excel(
        PROJECT_ROOT / "input" / "R_DMDT1.xlsx",
        aliases=merge_aliases("payable"),
        exact_phrase_overrides=overrides.get("payable", {}).get("exact_phrases", {}),
        supplemental_objects=overrides.get("payable", {}).get("supplemental_objects", []),
        own_company=own_company,
    )
    internal = ObjectMatcher.from_excel(
        PROJECT_ROOT / "input" / "MA NOI BO CTY.xlsx",
        aliases=merge_aliases("internal"),
        exact_phrase_overrides=overrides.get("internal", {}).get("exact_phrases", {}),
        supplemental_objects=overrides.get("internal", {}).get("supplemental_objects", []),
        own_company=own_company,
    )
    return receivable, payable, internal, EntityExtractor(own_company)


def _process_real(description: str, debit: float = 0, credit: float = 0, bank: str = "ACB"):
    receivable, payable, internal, extractor = _real_processing_context()
    return process_transaction(
        _txn(description, debit=debit, credit=credit, bank=bank),
        _config(),
        _engine(),
        receivable,
        payable,
        internal_matcher=internal,
        entity_extractor=extractor,
    )


def test_parsers_read_real_acb_vcb_msb_and_skip_msb_totals():
    acb = ACBParser().parse(_statement_sample("5614249_SAOKE_TK_202604 (2).xlsx"))
    vcb = VCBParser().parse(_statement_sample("lich-su-giao-dich-tai-khoan VCB T4.26.xls"))
    msb_parser = MSBParser()
    msb = msb_parser.parse(_statement_sample("ReportIBSCorpAccountStatement_20260526165427.xlsx"))

    assert any(item.debit_amount > 0 for item in acb)
    assert any(item.credit_amount > 0 for item in acb)
    assert any(item.debit_amount > 0 for item in vcb)
    assert any(item.credit_amount > 0 for item in vcb)
    assert any(item.debit_amount > 0 for item in msb)
    assert any(item.credit_amount > 0 for item in msb)
    assert not any(item.original_row_index >= 118 for item in msb)
    assert msb_parser.skipped_row_count >= 3


def test_msb_parser_accepts_beneficiary_counterparty_headers():
    parser = MSBParser()
    column_map = parser._match_header(
        [
            "Ngay giao dich/Transaction Date",
            "So but toan/Reference No",
            "Nguoi thu huong/Beneficiary",
            "Dien giai/Transaction Description",
            "No/Debit",
            "Co/Credit",
        ]
    )

    assert column_map["counterparty_raw"] == 2


def test_bao_co_rules_use_specific_accounts_before_customer_receivable():
    customer = [{"code": "ABC", "name": "Cong ty ABC"}]
    assert _process("ABC THANH TOAN CONG NO HD 123", credit=100, counterparty="ABC", receivable=customer).credit_account == "131"
    assert _process("LAI NHAP VON", credit=100).credit_account == "515"
    assert _process("CREDIT INTEREST", credit=100).credit_account == "515"
    assert _process("TRA LAI TAI KHOAN", credit=100).credit_account == "515"
    advance_refund = _process("HOAN LAI TAM UNG", credit=100, counterparty="LE NGOC DUC")
    assert advance_refund.credit_account == "141"
    assert advance_refund.object_code == "DUC"
    vcb_fx = _process("BAN NGOAI TE TY GIA USD VND", credit=100, bank="VCB")
    assert vcb_fx.credit_account == "1122VCB"
    assert vcb_fx.object_code == "VCB"
    msb_fx = _process("MUA TU BAO CO NGOAI TE THANH TOAN HD", credit=100, bank="MSB", receivable=customer)
    assert msb_fx.credit_account == "1122HB"
    assert msb_fx.object_code == "MSBHB"
    assert _process("GIAI NGAN KHOAN VAY", credit=100).credit_account == "34111"
    hoan_vay = _process("6097IBT1k1L92BMY-HOAN VAY", credit=100, bank="MSB", counterparty="9936363615/LE NGOC DUC")
    assert hoan_vay.status == "OK"
    assert hoan_vay.credit_account == "141"
    assert hoan_vay.object_code == "DUC"
    assert hoan_vay.matched_rule == "personal_advance_refund_to_company"
    fx1 = _process("M1HH/KHDN/ MUA TU BAO CO SO TIEN 42000 USD, TY GIA 26.217", credit=1101114000, bank="ACB")
    assert fx1.status == "OK"
    assert fx1.credit_account == "1122CT"
    assert fx1.object_code == "ACB"
    assert fx1.foreign_currency == "USD"
    assert fx1.foreign_amount == 42000
    assert fx1.exchange_rate == 26217
    assert fx1.reason == "Bán ngoại tệ 42000 USD tỷ giá 26217"
    fx2 = _process("M1HH/KHDN/ MUA TU BAO CO KH SO TIEN 50.000 USD, TY GIA 26228,", credit=1311400000, bank="ACB")
    assert fx2.status == "OK"
    assert fx2.credit_account == "1122CT"
    assert fx2.foreign_amount == 50000
    assert fx2.exchange_rate == 26228


def test_bao_no_rules_for_forex_loan_and_existing_bank_fee():
    vcb_buy = _process("MUA NGOAI TE", debit=100, bank="VCB")
    assert vcb_buy.debit_account == "1122VCB"
    assert vcb_buy.object_code == "VCB"
    msb_buy = _process("MUA USD", debit=100, bank="MSB")
    assert msb_buy.debit_account == "1122HB"
    assert msb_buy.object_code == "MSBHB"
    assert _process("THU NO TK VAY 001065887769", debit=100).debit_account == "34111"
    assert _process("TRA GOC KHOAN VAY", debit=100).debit_account == "34111"
    assert not _process("THANH TOAN USD", debit=100).debit_account.startswith("1122")
    vcb_fee = _process("PHI NGAN HANG", debit=100, bank="VCB")
    assert vcb_fee.debit_account == "635"
    assert vcb_fee.object_code == "VCB"
    acb_fee = _process("THU PHI PHAT HANH BAO LANH THUC HIEN HOP DONG", debit=100, bank="ACB")
    assert acb_fee.debit_account == "635"
    assert acb_fee.object_code == "ACB"
    msb_fee = _process("PHI CHUYEN TIEN", debit=100, bank="MSB")
    assert msb_fee.debit_account == "635"
    assert msb_fee.object_code == "MSBHB"
    assert _process("NOP THUE GTGT THANG 4", debit=100).debit_account == "3331"


def test_loan_account_repayment_and_customs_rules_use_templates():
    loan = _process("IBBIZ.6067857860.Thanh toan TK vay 00....7769", debit=100, bank="VCB")
    assert loan.status == "OK"
    assert loan.debit_account == "34111"
    assert loan.object_code == "VCB"
    assert loan.reason == "TT tài khoản vay 00...7769"
    assert loan.matched_rule == "loan_account_repayment"

    tkv = _process("TRANSFERTT TKV 1065887769", debit=100, bank="VCB")
    assert tkv.status == "OK"
    assert tkv.debit_account == "34111"
    assert tkv.reason == "TT tài khoản vay 00...7769"

    export_tax = _process(
        "2606226868048050.MS0200410388;Ch554;HQ03YY;LHB11;TK30865926994;NTK20062026;Thue;TM1851(XK);ST28625023",
        debit=28625023,
        bank="VCB",
    )
    assert export_tax.status == "OK"
    assert export_tax.debit_account == "3333"
    assert export_tax.object_code == "HQHP"
    assert export_tax.reason == "Nộp thuế xuất khẩu theo tờ khai: 30865926994"

    customs_fee = _process(
        "2606306868053212.MS0200410388;Ch554;HQ03YY;LHB11;TK30869474793;NTK30062026;;TM2663(LP);ST20000",
        debit=20000,
        bank="VCB",
    )
    assert customs_fee.status == "OK"
    assert customs_fee.debit_account == "64211"
    assert customs_fee.object_code == "HQHP"
    assert customs_fee.reason == "TT phí hải quan tờ khai xuất khẩu"

    port_infra = _process(
        "HQTP247.ID_CT:202620808391;LP:PHT01;DVNP:0200410388;DVTP:31;MA_CQT:1109448;TM:2267;ST:3000000",
        debit=3000000,
        bank="VCB",
    )
    assert port_infra.status == "OK"
    assert port_infra.debit_account == "331"
    assert port_infra.object_code == "CVDUONGTHUY"
    assert port_infra.reason == "TT phí hạ tầng công nghệ"

    customs_fee_other_amount = _process(
        "2606046868037169.MS0200410388;Ch554;HQ47NM;LHB11;TK9999999999-5.;NTK28052026;;TM2663(LP);ST240000",
        debit=240000,
        bank="VCB",
    )
    assert customs_fee_other_amount.status == "OK"
    assert customs_fee_other_amount.debit_account == "64211"
    assert customs_fee_other_amount.object_code == "HQHP"
    assert customs_fee_other_amount.reason == "TT phí hải quan tờ khai xuất khẩu"


def test_utility_rules_use_default_objects_and_reasons():
    electricity = _process("THANH TOAN TIEN DIEN KY HOA DON THANG 4", debit=100, bank="VCB")
    assert electricity.status == "OK"
    assert electricity.debit_account == "331"
    assert electricity.object_code == "DLUC"
    assert electricity.reason == "TT tiền điện"

    water = _process("HOA DON TIEN NUOC THANG 4", debit=100, bank="ACB")
    assert water.status == "OK"
    assert water.debit_account == "331"
    assert water.object_code == "CAPNUOC"
    assert water.reason == "TT tiền nước"


def test_fuel_payment_petrolimex_uses_petro_object_code():
    result = _process_real("CT TNHH LE PHAM TT TIEN XANG DAU PETROLIMEX", debit=100)

    assert result.status == "OK"
    assert result.debit_account == "331"
    assert result.object_code == "PETRO"
    assert result.reason.startswith("TT tiền xăng dầu")


def test_aaa_insurance_uses_specific_object_code_and_reason():
    debit = _process_real("CT TNHH LE PHAM THANH TOAN PHI BAO HIEM AAA", debit=100)
    assert debit.status == "OK"
    assert debit.flow == FLOW_BAO_NO
    assert debit.debit_account == "331"
    assert debit.object_code == "BAOHIEMAAA"
    assert debit.object_name
    assert debit.reason == "TT phí bảo hiểm AAA"

    credit = _process_real("BAO HIEM AAA HP THANH TOAN PHI BAO HIEM", credit=100)
    assert credit.status == "OK"
    assert credit.flow == FLOW_BAO_CO
    assert credit.credit_account == "131"
    assert credit.object_code == "BAOHIEMAAA"
    assert credit.object_name
    assert credit.reason == "TT phí bảo hiểm AAA"


def test_freight_reason_format_is_shared_for_bao_no_and_bao_co():
    records = [{"code": "VINHLONG", "name": "Công ty TNHH Vĩnh Long"}]
    payable = _process(
        "CT TNHH LE PHAM TT CUOC VAN CHUYEN CHO VINH LONG",
        debit=100,
        payable=records,
    )
    assert payable.status == "OK"
    assert payable.flow == FLOW_BAO_NO
    assert payable.object_code == "VINHLONG"
    assert payable.reason == "TT cước vận chuyển (Công ty TNHH Vĩnh Long)"

    receivable = _process("VINH LONG TT CUOC VAN CHUYEN", credit=100, receivable=records)
    assert receivable.status == "OK"
    assert receivable.flow == FLOW_BAO_CO
    assert receivable.object_code == "VINHLONG"
    assert receivable.reason == "TT cước vận chuyển (Công ty TNHH Vĩnh Long)"


def test_pil_vietnam_uses_s5_and_fixed_crew_change_reason():
    result = _process_real("PIL VIETNAM CO LTD PIL PAY INV 584", credit=100)

    assert result.status == "OK"
    assert result.flow == FLOW_BAO_CO
    assert result.object_code == "S5 VIET NAM"
    assert result.reason == "TT phí dịch vụ thay đổi thuyền viên (Chi nhánh công ty TNHH S5 Việt Nam tại Hải Phòng)"


def test_taubien_saigon_uses_vessel_in_reason():
    result = _process_real(
        "CTY VT VA DL TAU BIEN SAI GON TT TIEN TAU TAO TREASURE CHO CTY LE PHAM",
        credit=100,
    )

    assert result.status == "OK"
    assert result.flow == FLOW_BAO_CO
    assert result.object_code == "TAUBIENSAIGON"
    assert result.entities.vessel == "TAO TREASURE"
    assert result.reason == "TT tiền tàu TAO TREASURE (Công ty TNHH Vận Tải & Đại Lý Tàu Biển Sài Gòn)"


def test_tax_and_bank_loan_interest_default_objects():
    assert _process("NOP THUE GTGT THANG 4", debit=100, bank="ACB").object_code == "CUCTHUE"
    assert _process("NOP THUE TNDN TAM NOP", debit=100, bank="VCB").object_code == "CUCTHUE"
    assert _process("NOP THUE TNCN", debit=100, bank="MSB").object_code == "CUCTHUE"
    xnk = _process("NOP THUE HAI QUAN THUE XNK", debit=100, bank="VCB")
    assert xnk.debit_account == "3333"
    assert xnk.object_code == "CUCTHUE"
    import_vat = _process(
        "2604136868002202.MS0200410388;Ch554;HQ01B1;LHA11;TK10814236393;NTK13042026;Thue;TM1702(VA);ST9234971",
        debit=9234971,
        bank="VCB",
    )
    assert import_vat.status == "OK"
    assert import_vat.debit_account == "33312"
    assert import_vat.object_code == "HQHP"
    assert import_vat.reason == "Nộp thuế GTGT hàng nhập khẩu theo tờ khai 10814236393"

    vcb_interest = _process("TRA LAI VAY", debit=100, bank="VCB")
    assert vcb_interest.status == "OK"
    assert vcb_interest.debit_account == "635"
    assert vcb_interest.object_code == "VCB"
    assert vcb_interest.reason == "Lãi vay ngân hàng"

    acb_interest = _process("TRA LAI VAY", debit=100, bank="ACB")
    assert acb_interest.status == "ERROR"
    assert acb_interest.object_code != "ACB"


def test_customer_refund_out_uses_receivable_catalog():
    result = _process(
        "IBBIZ.6062736612.6110BFTVG231VAJ7.CT TNHH LE PHAM HOAN TIEN CHO CT TAN BINH TAU TAN BINH 234",
        debit=9160233,
        bank="VCB",
        receivable=[{"code": "TANBINH", "name": "Công ty TNHH Tân Bình"}],
    )

    assert result.status == "OK"
    assert result.flow == FLOW_BAO_NO
    assert result.debit_account == "131"
    assert result.credit_account == "1121VCB"
    assert result.object_code == "TANBINH"
    assert result.matched_rule == "customer_refund_out"


def test_acb_exception_patterns_pass_with_approved_overrides():
    receivable_cases = [
        ("CTY NGUYEN KIM 0109912477 STSTMSHP 2604109 GD 6125IBT1FJQI1KNQ", "NGUYENKIM"),
        ("PIL VIETNAM CO LTD TAX CODE - 0303449450-[3186493995] NHTMCP A CHAU HCM HCM PIL PAY INV 584", "S5 VIET NAM"),
        ("CTY TNHH PTXD VA TM 0101101999 PHI PHAT LENH SO BL STSTMSHP2604102", "PTXDVATM"),
        ("TCL SMART DEVICE VIET NAM COMPANY LIMITED-TIEN COC", "THONGMINHTCL"),
        ("CONG TY MINH HUY THANH TON TIEN THUE VAN PHONG VA DIEN NUOC THANG 02 03", "MINHHUY"),
        ("MBVCB.14166421961.394500.CTY SAKAI NOP PHI BILL 2603107", "SAKAI"),
        ("CP VIETNAM CORPORATION-553 VOI HOT-TA 25KGBAO CHARGEDETAILS OUR", "CPVIETNAM"),
        ("CN HN-CT TNHH MINH PHONG HOP NHAT CHUYEN TIEN DIEN NUOC T3/2026", "MINHPHONGHOPNHAT"),
        ("CO SO THANH PHONG CHUYEN TIEN GD 6135IBT1CJ1HNHAG", "THANHPHONG"),
        ("CHI NHANH CONG TY TNHH DO SUNG MACHINERY CHUYEN TIEN MUA KET SAT CHO CTY LE PHAM", "DOSUNG"),
    ]
    for description, object_code in receivable_cases:
        result = _process_real(description, credit=100)
        assert result.status == "OK"
        assert result.flow == FLOW_BAO_CO
        assert result.credit_account == "131"
        assert result.object_code == object_code

    payable_cases = [
        ("CT TNHH LE PHAM TT TIEN HANG CHO CT VIET THANG", "VIETTHAG"),
        ("LE PHAM CT CHO THIEN SON", "THIENSON"),
        ("LE PHAM CT CHO VOI VIET", "VOIVIET"),
        ("LE PHAM CT CHO 268", "26868"),
        ("TT TIENHANG CHO CT CP HC MINH DUC", "MD"),
        ("CT TNHH LE PHAM TT PHI GIAM DINH CHO SGS VIET NAM", "SGS"),
        ("LE PHAM DAT COC THIET KE PHAN MEM TU DONG HOA NGHIEP VU", "PHANMEM"),
        ("CTY LEPHAM TT CUOC BIEN CHO VSICO", "VSICO"),
        ("LE PHAM CT CHO NHAT MINH", "VTBNHATMINH"),
        ("CT TNHH LE PHAM TT CANG GAMA. TAU EAGLE AROW, ST 22376.67, TG 26125.", "CANGGAMA"),
    ]
    for description, object_code in payable_cases:
        result = _process_real(description, debit=100)
        assert result.status == "OK"
        assert result.flow == FLOW_BAO_NO
        assert result.debit_account == "331"
        assert result.object_code == object_code


def test_real_acb_june_configured_abbreviation_objects_do_not_fall_to_exception():
    payable_cases = [
        ("CT TNHH LE PHAM TT CHO CVU HH BINH THUAN . CONG NO THEO BANG KE DINH KEM", "CVHH-BINHTHUAN"),
        ("CTY LE PHAM TT CHO CVU THANH HOA. CONG NO THEO BANG KE.", "CVHH-THANHHOA"),
        ("CT TNHH LE PHAM TT CHO CT TM VT NHAT MINH HD 554", "VATTUNHATMINH"),
        ("CTY LE PHAM TT CHO CANG CONG TE NO QT CAI LAN. TAU QI YUAN 17", "QTCAILAN"),
        ("CTY LE PHAM TT CHO TT KS BENH TAT THANH HOA. CONG NO THEO BANG KE", "KDYT-TH"),
        ("CT TNHH LE PHAM TT CHO CTCP TD LOGISTICS QT HD 240", "LOGISTICS QUOCTE"),
        ("LE PHAM THANH TOAN TIEN BAO DUONG XE OTO CHO TOYOTA", "TOYOTA"),
    ]
    for description, object_code in payable_cases:
        result = _process_real(description, debit=100)
        assert result.status == "OK"
        assert result.object_code == object_code

    receivable_cases = [
        ("CTY BONG SEN TT TIEN DIEN, NUOC CHO CTY TNHH LE PHAM; HD: 892", "BONGSEN"),
        ("MBVCB.14903541098.340878.CTY BINH MINH, STSTMSHP2606109", "BINHMINH"),
        ("CTY CP DAO TAO VA CUNG UNG NHAN LUC VOS TT TIEN DICH VU THAY THUYEN VIEN", "NHANLUCVOSCO"),
    ]
    for description, object_code in receivable_cases:
        result = _process_real(description, credit=100)
        assert result.status == "OK"
        assert result.object_code == object_code


def test_bill_issue_fee_receipts_use_customer_001_and_company_reason():
    simple_fee = _process("KHACH HANG THANH TOAN PHI CAP LENH BILL", credit=100, bank="VCB")
    assert simple_fee.status == "OK"
    assert simple_fee.flow == FLOW_BAO_CO
    assert simple_fee.credit_account == "131"
    assert simple_fee.object_code == "001"
    assert simple_fee.reason == "TT phí cấp lệnh (Khách lẻ)"

    cases = [
        (
            "CTY TNHH PTXD VA TM 0101101999 PHI PHAT LENH SO BL STSTMSHP2604102",
            "CTY TNHH PTXD VA TM",
        ),
        (
            "MBVCB.14166421961.394500.CTY SAKAI NOP PHI BILL 2603107",
            "CTY SAKAI",
        ),
    ]
    for description, company_name in cases:
        result = _process_real(description, credit=1100000, bank="MSB")
        assert result.status == "OK"
        assert result.flow == FLOW_BAO_CO
        assert result.credit_account == "131"
        assert result.object_code == "001"
        assert result.object_name == company_name
        assert result.reason == "Thu phí cấp lệnh khách lẻ"


def test_bill_issue_fee_receipt_requires_amount_guard():
    result = _process_real(
        "CTY CLIO SHIPPING TT PHI DO BL 106043ZJ001 DEN ZJ009 MV LILA NOLA",
        credit=4860000,
    )
    assert result.status == "OK"
    assert result.object_code == "CLIO SHIPPING"

    same_amount_without_fee_signal = _process_real(
        "CTY NGUYEN KIM 0109912477 STSTMSHP 2604109 GD 6125IBT1FJQI1KNQ",
        credit=1100000,
    )
    assert same_amount_without_fee_signal.status == "OK"
    assert same_amount_without_fee_signal.object_code == "NGUYENKIM"


def test_company_advance_fee_keeps_internal_advance_person():
    advance = _process_real("T UNG CHO HOANG ANH TT PHI NANG HA CHO CT CP CANG HAI PHONG", debit=100)
    assert advance.status == "OK"
    assert advance.debit_account == "141"
    assert advance.object_code == "HOANGANH"


def test_acb_negative_patterns_stay_manual_review():
    assert _process_real("CHUYEN TIEN GD 6140IBT1FJW5PHSH", credit=100).status == "ERROR"
    own_transfer = _process_real("CONG TY TNHH LE PHAM CHUYEN TK GD 6141IBT1FJW4DKQJ", credit=100)
    assert own_transfer.status == "ERROR"
    assert own_transfer.credit_account != "131"


def test_cash_receipt_and_cash_payment_are_exclusive_flows():
    cash_in = _process("RUT TIEN MAT NHAP QUY", debit=100, bank="ACB")
    assert cash_in.flow == FLOW_THU_TIEN_MAT
    assert cash_in.debit_account == "1111"
    assert cash_in.credit_account == "1121CT"

    cheque_cash_in = _process("LE THI THANH HOA#001178043963#CHI SEC 22541668#1156992 ;", debit=3000000000, bank="ACB")
    assert cheque_cash_in.flow == FLOW_THU_TIEN_MAT
    assert cheque_cash_in.debit_account == "1111"
    assert cheque_cash_in.credit_account == "1121CT"
    assert cheque_cash_in.entities.cash_person_name == "LE THI THANH HOA"

    fee = _process("PHI RUT TIEN", debit=100, bank="VCB")
    assert fee.flow != FLOW_THU_TIEN_MAT

    cash_out = _process("NOP TIEN MAT VAO TAI KHOAN", credit=100, bank="MSB")
    assert cash_out.flow == FLOW_CHI_TIEN_MAT
    assert cash_out.debit_account == "1121HB"
    assert cash_out.credit_account == "1111"

    acb_cash_out = _process("LE THI THANH HOA#001178043963#NT#12739876;12739876-NT-TK-ACB-1794027", credit=100, bank="ACB")
    assert acb_cash_out.flow == FLOW_CHI_TIEN_MAT
    assert acb_cash_out.debit_account == "1121CT"
    assert acb_cash_out.credit_account == "1111"
    assert acb_cash_out.entities.cash_person_name == "LE THI THANH HOA"

    customer = [{"code": "KHACHHANG", "name": "Khach Hang"}]
    receivable = _process(
        "KHACH HANG NOP TIEN THANH TOAN HOA DON",
        credit=100,
        counterparty="KHACH HANG",
        receivable=customer,
    )
    assert receivable.flow == FLOW_BAO_CO
    assert receivable.credit_account == "131"


def test_personal_advance_to_company_goes_bao_co_141_not_receivable():
    cases = [
        "LE NGOC DUC CHUYEN TIEN GD 6124IBT1AWZTW1E7 040526-14:04:19",
        "LE NGOC DUC CHUYEN TIEN GD 6124IBT1AWZTW2S9 040526-14:04:55",
        "IB LE NGOC DUC CHUYEN KHOAN",
    ]
    for description in cases:
        result = _process(description, credit=100, bank="ACB")
        assert result.status == "OK"
        assert result.flow == FLOW_BAO_CO
        assert result.debit_account == "1121CT"
        assert result.credit_account == "141"
        assert result.object_code == "DUC"
        assert result.reason == "Nhận tiền tạm ứng cá nhân DUC"
        assert result.matched_rule == "personal_advance_to_company"


def test_advance_splits_company_payable_and_internal_person():
    company = _process(
        "TAM UNG CHO CONG TY ABC",
        debit=100,
        bank="ACB",
        payable=[{"code": "ABC", "name": "Cong ty ABC"}],
    )
    assert company.status == "OK"
    assert company.flow == FLOW_BAO_NO
    assert company.debit_account == "331"
    assert company.object_code == "ABC"
    assert company.reason == "Tạm ứng tiền hàng (Cong ty ABC)"

    person = _process("TAM UNG CHO LE NGOC DUC", debit=100, bank="ACB")
    assert person.status == "OK"
    assert person.flow == FLOW_BAO_NO
    assert person.debit_account == "141"
    assert person.object_code == "DUC"
    assert person.reason == "Tạm ứng cá nhân DUC"

    alias_person = _process("TAM UNG CHO VIET HUNG", debit=100, bank="ACB")
    assert alias_person.status == "OK"
    assert alias_person.flow == FLOW_BAO_NO
    assert alias_person.debit_account == "141"
    assert alias_person.object_code == "VIETHUNG"


def test_internal_advance_hoang_anh_uses_hoanganh_not_hoang():
    result = _process_real("TAM UNG CHO HOANG ANH", debit=12960000, bank="ACB")

    assert result.status == "OK"
    assert result.flow == FLOW_BAO_NO
    assert result.debit_account == "141"
    assert result.credit_account == "1121CT"
    assert result.object_code == "HOANGANH"
    assert result.reason == "Tạm ứng cá nhân HOANGANH"


def test_rule_first_does_not_call_ml_when_rule_matched():
    class RaisingClassifier:
        def predict(self, *args, **kwargs):  # noqa: ANN002, ANN003
            raise AssertionError("ML should not be called after a rule match")

    result = process_transaction(
        _txn("TRA LAI TAI KHOAN", credit=100),
        _config(),
        _engine(),
        ObjectMatcher([]),
        ObjectMatcher([]),
        classifier=RaisingClassifier(),
    )
    assert result.status == "OK"
    assert result.flow == FLOW_BAO_CO
    assert result.credit_account == "515"


def test_output_has_four_pad_sheets_with_exchange_rate_column(tmp_path):
    items = [
        _process("RUT TIEN MAT NHAP QUY", debit=100, bank="ACB"),
        _process("NOP TIEN MAT VAO TAI KHOAN", credit=200, bank="VCB"),
        _process("MUA USD", debit=300, bank="MSB"),
        _process("TRA LAI TAI KHOAN", credit=400, bank="MSB"),
    ]
    output_file = tmp_path / "rpa_input.xlsx"
    write_excel(items, output_file, run_id="run1", run_stats={"skipped_non_transaction_rows": 3})
    wb = load_workbook(output_file, data_only=True)

    for sheet_name in ["BAO_NO_INPUT", "BAO_CO_INPUT", "THU_TIEN_MAT_INPUT", "CHI_TIEN_MAT_INPUT"]:
        assert sheet_name in wb.sheetnames
        if sheet_name == "THU_TIEN_MAT_INPUT":
            expected_columns = RPA_THU_TIEN_MAT_COLUMNS
        elif sheet_name == "CHI_TIEN_MAT_INPUT":
            expected_columns = RPA_CHI_TIEN_MAT_COLUMNS
        else:
            expected_columns = RPA_BUSINESS_COLUMNS
        assert [cell.value for cell in wb[sheet_name][1]] == expected_columns
        assert wb[sheet_name].max_column == len(expected_columns)
    assert "RPA_TASKS" in wb.sheetnames
    assert "AUDIT_LOG" in wb.sheetnames
    assert "MANUAL_REVIEW" in wb.sheetnames


def test_rpa_input_exports_foreign_exchange_rate(tmp_path):
    bao_co = _process(
        "M1HH/KHDN/ MUA TU BAO CO SO TIEN 42000 USD, TY GIA 26.217",
        credit=1101114000,
        bank="ACB",
    )
    bao_no = _process(
        "MUA NGOAI TE SO TIEN 1000 USD, TY GIA 25.100",
        debit=25100000,
        bank="ACB",
    )
    output_file = tmp_path / "rpa_input.xlsx"
    write_excel([bao_co, bao_no], output_file)
    wb = load_workbook(output_file, data_only=True)

    bao_co_headers = [cell.value for cell in wb["BAO_CO_INPUT"][1]]
    bao_co_values = dict(zip(bao_co_headers, [cell.value for cell in wb["BAO_CO_INPUT"][2]]))
    assert bao_co_values["Tỷ giá"] == 26217

    bao_no_headers = [cell.value for cell in wb["BAO_NO_INPUT"][1]]
    bao_no_values = dict(zip(bao_no_headers, [cell.value for cell in wb["BAO_NO_INPUT"][2]]))
    assert bao_no_values["Tỷ giá"] == 25100


def test_cash_flows_export_recipient_name(tmp_path):
    cash_in = _process("LE THI THANH HOA#001178043963#CHI SEC 22541668#1156992 ;", debit=3000000000, bank="ACB")
    cash_out = _process("LE THI THANH HOA#001178043963#NT#12739876;12739876-NT-TK-ACB-1794027", credit=100, bank="ACB")
    output_file = tmp_path / "rpa_input.xlsx"
    write_excel([cash_in, cash_out], output_file)
    wb = load_workbook(output_file, data_only=True)

    thu_headers = [cell.value for cell in wb["THU_TIEN_MAT_INPUT"][1]]
    thu_values = dict(zip(thu_headers, [cell.value for cell in wb["THU_TIEN_MAT_INPUT"][2]]))
    assert thu_values["Người nhận tiền"] == "LE THI THANH HOA"

    chi_headers = [cell.value for cell in wb["CHI_TIEN_MAT_INPUT"][1]]
    chi_values = dict(zip(chi_headers, [cell.value for cell in wb["CHI_TIEN_MAT_INPUT"][2]]))
    assert chi_values["Người nộp tiền"] == "LE THI THANH HOA"


def test_integration_process_real_samples_and_write_outputs(tmp_path):
    import logging

    config = _config()
    logger = logging.getLogger("integration-four-flows")
    processed = process_all(
        statements_dir=STATEMENTS_DIR,
        receivable_path=PROJECT_ROOT / "input" / "R_DMDT1 1.xlsx",
        payable_path=PROJECT_ROOT / "input" / "R_DMDT1.xlsx",
        rules_path=None,
        default_rules_path=PROJECT_ROOT / "config" / "default_rules.yaml",
        config=config,
        logger=logger,
    )
    result = write_outputs(processed, tmp_path, config)
    wb = load_workbook(result.excel_path, data_only=True)

    assert len(processed) > 0
    assert not any(item.bank == "MSB" and item.original_row_index >= 118 for item in processed)
    for sheet_name in ["BAO_NO_INPUT", "BAO_CO_INPUT", "THU_TIEN_MAT_INPUT", "CHI_TIEN_MAT_INPUT"]:
        assert sheet_name in wb.sheetnames
        if sheet_name == "THU_TIEN_MAT_INPUT":
            expected_columns = RPA_THU_TIEN_MAT_COLUMNS
        elif sheet_name == "CHI_TIEN_MAT_INPUT":
            expected_columns = RPA_CHI_TIEN_MAT_COLUMNS
        else:
            expected_columns = RPA_BUSINESS_COLUMNS
        if config.get("output", {}).get("rpa_reason_encoding") == "tcvn3":
            expected_columns = list(expected_columns)
            expected_columns.insert(expected_columns.index("Lí do") + 1, RPA_REASON_UNICODE_COLUMN)
        assert [cell.value for cell in wb[sheet_name][1]] == expected_columns
    assert "SUMMARY" in wb.sheetnames
    assert "RPA_TASKS" in wb.sheetnames
