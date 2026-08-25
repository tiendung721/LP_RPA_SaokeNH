from __future__ import annotations

from datetime import date

from openpyxl import load_workbook

from src.flows import FLOW_BAO_NO
from src.models import ProcessedTransaction
from src.output_writer import write_excel
from src.rpa_input_sanitizer import sanitize_rpa_input
from src.rpa_input_status import TRANSACTION_UID_COLUMN


def _processed(uid: str, amount: int) -> ProcessedTransaction:
    return ProcessedTransaction(
        source_file="sample.xlsx",
        original_row_index=2,
        bank="ACB",
        flow=FLOW_BAO_NO,
        transaction_date=date(2026, 6, 20),
        object_code="ABC",
        object_name="ABC",
        reason="Thanh toan ABC",
        debit_account="331",
        credit_account="1121CT",
        amount=amount,
        use_case="Thanh toan",
        original_content=f"TT ABC {uid}",
        counterparty_raw="ABC",
        doc_no=uid,
        status="OK",
        error_note="",
        confidence=0.95,
        transaction_uid=uid,
        source_sheet="Statement",
    )


def test_sanitize_rpa_input_deletes_blank_rows_without_legacy_tasks_sheet(tmp_path):
    input_file = tmp_path / "rpa_input.xlsx"
    write_excel([_processed("uid_skip", 100), _processed("uid_keep", 200)], input_file, run_id="run1")

    workbook = load_workbook(input_file)
    ws = workbook["BAO_NO_INPUT"]
    for column_index in range(1, ws.max_column + 1):
        ws.cell(row=2, column=column_index).value = None
    workbook.save(input_file)

    summary = sanitize_rpa_input(input_file)

    assert summary.saved is True
    assert summary.removed_rows == [{"sheet": "BAO_NO_INPUT", "row": 2, "reason": "blank"}]
    assert summary.removed_tasks == []
    assert summary.updated_tasks == 0

    workbook = load_workbook(input_file, data_only=True)
    ws = workbook["BAO_NO_INPUT"]
    headers = [cell.value for cell in ws[1]]
    uid_column = headers.index(TRANSACTION_UID_COLUMN) + 1
    assert ws.max_row == 2
    assert ws.cell(row=2, column=uid_column).value == "uid_keep"

    assert "RPA_TASKS" not in workbook.sheetnames


def test_sanitize_rpa_input_removes_rows_missing_transaction_uid(tmp_path):
    input_file = tmp_path / "rpa_input.xlsx"
    write_excel([_processed("uid_missing_later", 100)], input_file, run_id="run1")

    workbook = load_workbook(input_file)
    ws = workbook["BAO_NO_INPUT"]
    headers = [cell.value for cell in ws[1]]
    uid_column = headers.index(TRANSACTION_UID_COLUMN) + 1
    ws.cell(row=2, column=uid_column, value="")
    workbook.save(input_file)

    summary = sanitize_rpa_input(input_file)

    assert summary.removed_rows == [{"sheet": "BAO_NO_INPUT", "row": 2, "reason": "missing_transaction_uid"}]
    assert summary.removed_tasks == []
    assert summary.updated_tasks == 0

    workbook = load_workbook(input_file, data_only=True)
    assert workbook["BAO_NO_INPUT"].max_row == 1
    assert "RPA_TASKS" not in workbook.sheetnames
