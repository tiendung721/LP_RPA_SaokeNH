from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.excel_io import atomic_output_path
from src.normalizer import clean_display_text, normalize_text
from src.object_matcher import load_catalog
from src.vietnamese_encoding import unicode_to_tcvn3

from .models import CATALOG_DEFINITIONS, ManagedObject
from .paths import RuleManagerPaths


class CatalogWorkbookError(RuntimeError):
    """Raised when a VACOM catalog workbook cannot be updated safely."""


class ObjectCatalogStore:
    def __init__(self, paths: RuleManagerPaths):
        self.paths = paths

    def list_objects(self, catalog: str) -> list[ManagedObject]:
        self._definition(catalog)
        path = self.paths.catalog_path(catalog)
        objects = load_catalog(path)
        return [
            ManagedObject(
                catalog=catalog,
                code=item.code,
                name=clean_display_text(item.name),
                tax_code=item.tax_code,
                address=clean_display_text(item.address),
                group_code=item.group_code,
            )
            for item in objects
        ]

    def find(self, catalog: str, code: str) -> ManagedObject | None:
        normalized_code = normalize_text(code)
        for item in self.list_objects(catalog):
            if normalize_text(item.code) == normalized_code:
                return item
        return None

    def append(self, item: ManagedObject) -> int:
        definition = self._definition(item.catalog)
        path = self.paths.catalog_path(item.catalog)
        if not path.exists():
            raise FileNotFoundError(f"Không tìm thấy danh mục: {path}")
        if self.find(item.catalog, item.code):
            raise CatalogWorkbookError(f"Mã ĐT {item.code} đã tồn tại trong danh mục {definition.label}")

        workbook = load_workbook(path)
        sheet_name = "r_dmdt" if "r_dmdt" in workbook.sheetnames else workbook.sheetnames[0]
        ws = workbook[sheet_name]
        columns = self._detect_columns(ws)
        target_row = self._next_data_row(ws, columns["code"])
        source_row = target_row - 1
        self._copy_row_style(ws, source_row, target_row)

        values: dict[str, Any] = {
            "code": item.code.strip(),
            "sequence": 1,
            "name": self._encode_text(item.name, definition.tcvn3_text),
            "group_name": self._encode_text(definition.group_name, definition.tcvn3_text),
            "address": self._encode_text(item.address, definition.tcvn3_text),
            "phone": "",
            "tax_code": item.tax_code.strip(),
            "representative": "",
            "group_code": definition.group_code,
            "note": "",
            "unit": definition.unit,
        }
        for field, column_index in columns.items():
            if field in values:
                ws.cell(row=target_row, column=column_index, value=values[field])

        try:
            with atomic_output_path(path) as temporary_path:
                workbook.save(temporary_path)
                # Windows cannot atomically replace a workbook while openpyxl
                # still holds the source ZIP handle.
                workbook.close()
        finally:
            workbook.close()

        # Re-read through the same loader used by bank_agent.py. This catches a
        # structurally valid workbook that nevertheless cannot be used as a catalog.
        if not self.find(item.catalog, item.code):
            raise CatalogWorkbookError(
                f"Đã ghi workbook nhưng backend không đọc lại được Mã ĐT {item.code}"
            )
        return target_row

    @staticmethod
    def _definition(catalog: str):
        if catalog not in CATALOG_DEFINITIONS:
            raise ValueError(f"Danh mục không hỗ trợ: {catalog}")
        return CATALOG_DEFINITIONS[catalog]

    @staticmethod
    def _detect_columns(ws) -> dict[str, int]:
        columns: dict[str, int] = {}
        for row_index in range(1, min(ws.max_row, 20) + 1):
            for column_index in range(1, ws.max_column + 1):
                value = normalize_text(ws.cell(row=row_index, column=column_index).value)
                if not value:
                    continue
                if "code" not in columns and ("MA DT" in value or "MA DOI TUONG" in value):
                    columns["code"] = column_index
                elif "name" not in columns and ("TEN DOI TUONG" in value or value == "TEN DT"):
                    columns["name"] = column_index
                elif "group_name" not in columns and "TEN NHOM" in value:
                    columns["group_name"] = column_index
                elif "address" not in columns and "DIA CHI" in value:
                    columns["address"] = column_index
                elif "phone" not in columns and "DIEN THOAI" in value:
                    columns["phone"] = column_index
                elif "tax_code" not in columns and ("MS THUE" in value or "MA SO THUE" in value):
                    columns["tax_code"] = column_index
                elif "representative" not in columns and "NGUOI DAI DIEN" in value:
                    columns["representative"] = column_index
                elif "group_code" not in columns and "MA NHOM DT" in value:
                    columns["group_code"] = column_index
                elif "note" not in columns and "DIEN GIAI" in value:
                    columns["note"] = column_index
                elif "unit" not in columns and "DON VI" in value:
                    columns["unit"] = column_index

        if "code" not in columns or "name" not in columns:
            raise CatalogWorkbookError("Không nhận diện được cột Mã ĐT/Tên đối tượng trong workbook")
        # VACOM exports use an unlabeled sequence column immediately after Mã ĐT.
        columns.setdefault("sequence", columns["code"] + 1)
        return columns

    @staticmethod
    def _next_data_row(ws, code_column: int) -> int:
        last_data_row = 0
        for row_index in range(1, ws.max_row + 1):
            value = str(ws.cell(row=row_index, column=code_column).value or "").strip()
            if value and not (value.startswith("[") and value.endswith("]")):
                last_data_row = row_index
        if last_data_row < 1:
            raise CatalogWorkbookError("Không tìm thấy vùng dữ liệu danh mục")
        return last_data_row + 1

    @staticmethod
    def _copy_row_style(ws, source_row: int, target_row: int) -> None:
        if source_row < 1:
            return
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        for column_index in range(1, ws.max_column + 1):
            source = ws.cell(row=source_row, column=column_index)
            target = ws.cell(row=target_row, column=column_index)
            if source.has_style:
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
                target.alignment = copy(source.alignment)
                target.protection = copy(source.protection)
                target.number_format = source.number_format

    @staticmethod
    def _encode_text(value: str, use_tcvn3: bool) -> str:
        text = str(value or "").strip()
        return unicode_to_tcvn3(text) if use_tcvn3 else text
