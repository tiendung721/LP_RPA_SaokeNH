from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .excel_io import atomic_output_path, backup_legacy_workbook_once
from .flows import FLOW_BAO_CO, FLOW_BAO_NO, FLOW_CHI_TIEN_MAT, FLOW_THU_TIEN_MAT
from .models import ProcessedTransaction
from .rpa_tracking import (
    ELIGIBLE_RPA_STATUSES,
    ATTEMPT_SUCCESS,
    STATUS_DONE,
    STATUS_PENDING,
    abort_run_records,
    apply_status_update,
    finalize_run_records,
    normalize_status,
    reset_all_records,
    validate_status,
)
from .transaction_identity import assign_transaction_uids


SUMMARY_SHEET_NAME = "RPA_SUMMARY"
SUMMARY_VISIBLE_COLUMNS = [
    "Ngày CT",
    "Ngân hàng",
    "Luồng",
    "Nội dung giao dịch",
    "Mã ĐT",
    "Tên ĐT",
    "TK nợ",
    "TK có",
    "Thành tiền",
    "Kết quả phân loại",
    "Trạng thái RPA",
    "Số chứng từ VACOM",
    "Thông báo RPA",
    "Thời gian hoàn thành",
    "Nguồn sao kê",
]
SUMMARY_TECHNICAL_COLUMNS = [
    "transaction_uid",
    "last_run_id",
    "last_attempt_result",
    "rpa_started_at",
    "rpa_finished_at",
]
SUMMARY_COLUMNS = [*SUMMARY_VISIBLE_COLUMNS, *SUMMARY_TECHNICAL_COLUMNS]

STATUS_COLUMN = "Trạng thái RPA"
MESSAGE_COLUMN = "Thông báo RPA"
VOUCHER_COLUMN = "Số chứng từ VACOM"
COMPLETED_COLUMN = "Thời gian hoàn thành"
EDITABLE_COLUMNS = {STATUS_COLUMN, VOUCHER_COLUMN, MESSAGE_COLUMN}


@dataclass
class RpaRunState:
    run_id: str
    summary_df: pd.DataFrame
    rpa_items: list[ProcessedTransaction]
    stats: dict[str, int] = field(default_factory=dict)


def make_run_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def prepare_rpa_run(
    processed: list[ProcessedTransaction],
    summary_path: str | Path,
    tracking_path: str | Path | None = None,
    run_id: str | None = None,
    logger: Any | None = None,
) -> RpaRunState:
    run_id = run_id or make_run_id()
    if any(not item.transaction_uid for item in processed):
        assign_transaction_uids(processed)

    existing_df = load_summary(summary_path)
    existing_by_uid = {
        _clean_text(row.get("transaction_uid")): _clean_record(row)
        for row in existing_df.to_dict("records")
        if _clean_text(row.get("transaction_uid"))
    }
    records_by_uid: dict[str, dict[str, Any]] = {}
    rpa_items: list[ProcessedTransaction] = []
    seen: set[str] = set()
    stats = _initial_stats(processed)

    for item in processed:
        uid = item.transaction_uid
        if not uid:
            continue
        previous = existing_by_uid.get(uid)
        if previous is None:
            stats["new_count"] += 1
        record = _merge_record(previous, item, run_id)
        records_by_uid[uid] = record
        seen.add(uid)

        item.rpa_status = _clean_text(record.get(STATUS_COLUMN))
        item.rpa_message = _clean_text(record.get(MESSAGE_COLUMN))
        if item.rpa_status == STATUS_PENDING:
            stats["pending_count"] += 1
        if item.rpa_status == STATUS_DONE:
            stats["skipped_completed_count"] += 1
        if item.status == "OK" and item.rpa_status in ELIGIBLE_RPA_STATUSES:
            rpa_items.append(item)
            stats["auto_process_count"] += 1
            stats["waiting_count"] += 1
            _increment_flow_stat(stats, item.flow)

    for row in existing_df.to_dict("records"):
        uid = _clean_text(row.get("transaction_uid"))
        if uid and uid not in seen:
            records_by_uid[uid] = _ensure_summary_record(row)

    summary_df = _ensure_columns(pd.DataFrame(list(records_by_uid.values()), columns=SUMMARY_COLUMNS))
    return RpaRunState(run_id=run_id, summary_df=summary_df, rpa_items=rpa_items, stats=stats)


def load_summary(path: str | Path) -> pd.DataFrame:
    path = Path(path)
    if not path.exists():
        return pd.DataFrame(columns=SUMMARY_COLUMNS)
    try:
        df = pd.read_excel(path, sheet_name=SUMMARY_SHEET_NAME, dtype=object)
    except ValueError:
        df = pd.read_excel(path, sheet_name=0, dtype=object)
    records = [_ensure_summary_record(_clean_record(row)) for row in df.to_dict("records")]
    return pd.DataFrame(records, columns=SUMMARY_COLUMNS)


def write_summary(summary_df: pd.DataFrame, path: str | Path) -> None:
    path = Path(path)
    df = _ensure_columns(summary_df)
    legacy_schema = _has_legacy_schema(path)

    with atomic_output_path(path) as temporary_path:
        with pd.ExcelWriter(temporary_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=SUMMARY_SHEET_NAME, index=False)
            ws = writer.book[SUMMARY_SHEET_NAME]
            _format_sheet(ws)
            _apply_summary_controls(ws)
        if legacy_schema:
            backup_legacy_workbook_once(path, path.parent / "backup")


def update_rpa_status(
    summary_path: str | Path,
    transaction_uid: str,
    status: str,
    run_id: str = "",
    message: str = "",
    voucher_no: str = "",
) -> dict[str, Any]:
    status = validate_status(status)
    df = load_summary(summary_path)
    uid = _clean_text(transaction_uid)
    matches = df.index[df["transaction_uid"].astype(str) == uid].tolist()
    if not matches:
        raise KeyError(f"transaction_uid not found in RPA summary: {uid}")

    row_idx = matches[0]
    record = _clean_record(df.loc[row_idx].to_dict())
    updated_tracking = apply_status_update(
        _to_tracking_record(record),
        status,
        message=message,
        voucher_no=voucher_no,
        run_id=run_id,
    )
    updated = _apply_tracking_record(record, updated_tracking)
    for column in SUMMARY_COLUMNS:
        df.at[row_idx, column] = updated.get(column, "")
    write_summary(df, summary_path)
    return updated


def mark_rpa_started(summary_path: str | Path, transaction_uid: str, run_id: str) -> None:
    update_rpa_status(summary_path, transaction_uid, STATUS_PENDING, run_id=run_id)


def mark_rpa_done(
    summary_path: str | Path,
    transaction_uid: str,
    run_id: str,
    voucher_no: str = "",
    message: str = "",
) -> None:
    update_rpa_status(summary_path, transaction_uid, STATUS_DONE, run_id=run_id, message=message, voucher_no=voucher_no)


def mark_rpa_error(summary_path: str | Path, transaction_uid: str, run_id: str, message: str) -> None:
    update_rpa_status(summary_path, transaction_uid, STATUS_PENDING, run_id=run_id, message=message)


def finalize_rpa_run(summary_path: str | Path, run_id: str) -> pd.DataFrame:
    return _apply_bulk_tracking_update(summary_path, lambda records: finalize_run_records(records, run_id))


def abort_rpa_run(summary_path: str | Path, run_id: str, message: str = "") -> pd.DataFrame:
    return _apply_bulk_tracking_update(
        summary_path,
        lambda records: abort_run_records(records, run_id, message=message),
    )


def reset_all_rpa_status(summary_path: str | Path, message: str = "") -> pd.DataFrame:
    return _apply_bulk_tracking_update(
        summary_path,
        lambda records: reset_all_records(records, message=message),
    )


def _apply_bulk_tracking_update(summary_path: str | Path, updater) -> pd.DataFrame:
    df = load_summary(summary_path)
    records = [_ensure_summary_record(row) for row in df.to_dict("records")]
    tracking_by_uid = {
        _clean_text(record.get("transaction_uid")): record
        for record in updater([_to_tracking_record(record) for record in records])
    }
    updated = [
        _apply_tracking_record(record, tracking_by_uid.get(_clean_text(record.get("transaction_uid")), {}))
        for record in records
    ]
    result = pd.DataFrame(updated, columns=SUMMARY_COLUMNS)
    write_summary(result, summary_path)
    return result


def _merge_record(previous: dict[str, Any] | None, item: ProcessedTransaction, run_id: str) -> dict[str, Any]:
    current = _record_from_item(item, run_id)
    if not previous:
        current[STATUS_COLUMN] = STATUS_PENDING
        current[MESSAGE_COLUMN] = "" if item.status == "OK" else item.error_note
        return current

    previous = _ensure_summary_record(previous)
    previous_status = normalize_status(previous.get(STATUS_COLUMN), default=STATUS_PENDING)
    if previous_status == STATUS_DONE:
        return previous

    for field in (
        STATUS_COLUMN,
        VOUCHER_COLUMN,
        MESSAGE_COLUMN,
        COMPLETED_COLUMN,
        "last_attempt_result",
        "rpa_started_at",
        "rpa_finished_at",
    ):
        current[field] = previous.get(field, "")
    current[STATUS_COLUMN] = STATUS_PENDING
    current[MESSAGE_COLUMN] = item.error_note if item.status != "OK" else ""
    return current


def _record_from_item(item: ProcessedTransaction, run_id: str) -> dict[str, Any]:
    return {
        "Ngày CT": item.transaction_date,
        "Ngân hàng": item.bank,
        "Luồng": item.flow,
        "Nội dung giao dịch": item.original_content,
        "Mã ĐT": item.object_code,
        "Tên ĐT": item.object_name,
        "TK nợ": item.debit_account,
        "TK có": item.credit_account,
        "Thành tiền": item.amount,
        "Kết quả phân loại": item.status,
        STATUS_COLUMN: "",
        VOUCHER_COLUMN: "",
        MESSAGE_COLUMN: "",
        COMPLETED_COLUMN: "",
        "Nguồn sao kê": _source_display(item.source_file, item.source_sheet, item.original_row_index),
        "transaction_uid": item.transaction_uid,
        "last_run_id": run_id,
        "last_attempt_result": "",
        "rpa_started_at": "",
        "rpa_finished_at": "",
    }


def _ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    records = [_ensure_summary_record(_clean_record(row)) for row in df.to_dict("records")]
    return pd.DataFrame(records, columns=SUMMARY_COLUMNS)


def _ensure_summary_record(row: dict[str, Any]) -> dict[str, Any]:
    if any(column in row for column in SUMMARY_VISIBLE_COLUMNS):
        record = {column: _clean_cell(row.get(column, "")) for column in SUMMARY_COLUMNS}
    else:
        record = _migrate_legacy_record(row)

    status = normalize_status(
        record.get(STATUS_COLUMN) or row.get("rpa_status") or row.get("status"),
        default=STATUS_PENDING,
    )
    if status != STATUS_DONE and _clean_text(record.get("last_attempt_result")) == ATTEMPT_SUCCESS:
        status = STATUS_DONE
        record["last_attempt_result"] = ""
    record[STATUS_COLUMN] = status
    if status == STATUS_DONE and not record.get(COMPLETED_COLUMN):
        record[COMPLETED_COLUMN] = (
            record.get("rpa_finished_at")
            or row.get("completed_at")
            or row.get("updated_at")
            or _now()
        )
    return {column: _clean_cell(record.get(column, "")) for column in SUMMARY_COLUMNS}


def _migrate_legacy_record(row: dict[str, Any]) -> dict[str, Any]:
    source_file = row.get("source_file", "")
    source_sheet = row.get("source_sheet", "")
    source_row = row.get("source_row") or row.get("source_row_index") or row.get("original_row_index") or ""
    return {
        "Ngày CT": row.get("transaction_date", ""),
        "Ngân hàng": row.get("bank") or row.get("bank_code") or "",
        "Luồng": row.get("flow") or row.get("direction") or "",
        "Nội dung giao dịch": row.get("original_content", ""),
        "Mã ĐT": row.get("object_code") or row.get("matched_object_code") or "",
        "Tên ĐT": row.get("object_name") or row.get("matched_object_name") or "",
        "TK nợ": row.get("debit_account", ""),
        "TK có": row.get("credit_account", ""),
        "Thành tiền": row.get("amount", ""),
        "Kết quả phân loại": row.get("processing_status", ""),
        STATUS_COLUMN: row.get("rpa_status") or row.get("status") or STATUS_PENDING,
        VOUCHER_COLUMN: row.get("voucher_no", ""),
        MESSAGE_COLUMN: row.get("rpa_message", ""),
        COMPLETED_COLUMN: row.get("completed_at", ""),
        "Nguồn sao kê": _source_display(source_file, source_sheet, source_row),
        "transaction_uid": row.get("transaction_uid", ""),
        "last_run_id": row.get("last_run_id", ""),
        "last_attempt_result": row.get("last_attempt_result", ""),
        "rpa_started_at": row.get("rpa_started_at", ""),
        "rpa_finished_at": row.get("rpa_finished_at", ""),
    }


def _to_tracking_record(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "transaction_uid": record.get("transaction_uid", ""),
        "rpa_status": record.get(STATUS_COLUMN, STATUS_PENDING),
        "status": record.get(STATUS_COLUMN, STATUS_PENDING),
        "rpa_message": record.get(MESSAGE_COLUMN, ""),
        "voucher_no": record.get(VOUCHER_COLUMN, ""),
        "completed_at": record.get(COMPLETED_COLUMN, ""),
        "last_run_id": record.get("last_run_id", ""),
        "last_attempt_result": record.get("last_attempt_result", ""),
        "rpa_started_at": record.get("rpa_started_at", ""),
        "rpa_finished_at": record.get("rpa_finished_at", ""),
    }


def _apply_tracking_record(record: dict[str, Any], tracking: dict[str, Any]) -> dict[str, Any]:
    result = _ensure_summary_record(record)
    if not tracking:
        return result
    result[STATUS_COLUMN] = normalize_status(tracking.get("rpa_status") or tracking.get("status"), default=STATUS_PENDING)
    result[MESSAGE_COLUMN] = _clean_cell(tracking.get("rpa_message", ""))
    result[VOUCHER_COLUMN] = _clean_cell(tracking.get("voucher_no", ""))
    result[COMPLETED_COLUMN] = _clean_cell(tracking.get("completed_at", ""))
    for column in ("last_run_id", "last_attempt_result", "rpa_started_at", "rpa_finished_at"):
        result[column] = _clean_cell(tracking.get(column, result.get(column, "")))
    return result


def _initial_stats(processed: list[ProcessedTransaction]) -> dict[str, int]:
    return {
        "total_processed": len(processed),
        "new_count": 0,
        "auto_process_count": 0,
        "pending_count": 0,
        "in_progress_count": 0,
        "error_count": 0,
        "skipped_count": 0,
        "skipped_completed_count": 0,
        "waiting_count": 0,
        "retry_error_count": 0,
        "review_count": 0,
        "exception_count": sum(1 for item in processed if item.status != "OK"),
        "bao_no_output_count": 0,
        "bao_co_output_count": 0,
        "thu_tien_mat_output_count": 0,
        "chi_tien_mat_output_count": 0,
    }


def _increment_flow_stat(stats: dict[str, int], flow: str) -> None:
    mapping = {
        FLOW_BAO_NO: "bao_no_output_count",
        FLOW_BAO_CO: "bao_co_output_count",
        FLOW_THU_TIEN_MAT: "thu_tien_mat_output_count",
        FLOW_CHI_TIEN_MAT: "chi_tien_mat_output_count",
    }
    key = mapping.get(flow)
    if key:
        stats[key] += 1


def _has_legacy_schema(path: Path) -> bool:
    if not path.exists():
        return False
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        ws = workbook[SUMMARY_SHEET_NAME] if SUMMARY_SHEET_NAME in workbook.sheetnames else workbook.worksheets[0]
        headers = {str(cell.value or "").strip() for cell in ws[1]}
        workbook.close()
    except (OSError, ValueError):
        return False
    return "transaction_date" in headers or "rpa_status" in headers or "STATUS_COUNTS" in headers


def _source_display(source_file: Any, source_sheet: Any, source_row: Any) -> str:
    parts = [_clean_text(source_file), _clean_text(source_sheet)]
    row = _clean_text(source_row)
    if row:
        parts.append(f"dòng {row}")
    return " | ".join(part for part in parts if part)


def _clean_record(row: dict[str, Any]) -> dict[str, Any]:
    return {str(key): _clean_cell(value) for key, value in row.items()}


def _clean_cell(value: Any) -> Any:
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return value


def _clean_text(value: Any) -> str:
    value = _clean_cell(value)
    return str(value or "").strip()


def _now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    for name, column_index in headers.items():
        if name in {"Ngày CT", COMPLETED_COLUMN, "rpa_started_at", "rpa_finished_at"}:
            for row_index in range(2, ws.max_row + 1):
                ws.cell(row=row_index, column=column_index).number_format = "DD/MM/YYYY HH:MM:SS"
        if name == "Thành tiền":
            for row_index in range(2, ws.max_row + 1):
                ws.cell(row=row_index, column=column_index).number_format = "#,##0"
        letter = get_column_letter(column_index)
        max_len = max((len(str(cell.value)) for cell in ws[letter] if cell.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)


def _apply_summary_controls(ws) -> None:
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    gray = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    for name, column_index in headers.items():
        fill = yellow if name in EDITABLE_COLUMNS else gray
        for row_index in range(2, ws.max_row + 1):
            ws.cell(row=row_index, column=column_index).fill = fill
    comments = {
        STATUS_COLUMN: "Chỉ dùng chua_nhap hoặc hoan_thanh.",
        VOUCHER_COLUMN: "Số chứng từ VACOM; có thể bổ sung thủ công.",
        MESSAGE_COLUMN: "Ghi chú lỗi hoặc lý do cần nhập lại.",
        "transaction_uid": "Mã định danh hệ thống, không chỉnh sửa.",
    }
    for name, message in comments.items():
        if name in headers:
            ws.cell(row=1, column=headers[name]).comment = Comment(message, "Bank Agent")

    status_column = headers[STATUS_COLUMN]
    status_letter = get_column_letter(status_column)
    validation = DataValidation(type="list", formula1='"chua_nhap,hoan_thanh"', allow_blank=False)
    validation.error = "Chỉ chọn chua_nhap hoặc hoan_thanh"
    ws.add_data_validation(validation)
    validation.add(f"{status_letter}2:{status_letter}1048576")
    status_range = f"{status_letter}2:{status_letter}{max(ws.max_row, 2)}"
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(
            operator="equal",
            formula=['"chua_nhap"'],
            fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        ),
    )
    ws.conditional_formatting.add(
        status_range,
        CellIsRule(
            operator="equal",
            formula=['"hoan_thanh"'],
            fill=PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),
        ),
    )
    for column in SUMMARY_TECHNICAL_COLUMNS:
        ws.column_dimensions[get_column_letter(headers[column])].hidden = True
