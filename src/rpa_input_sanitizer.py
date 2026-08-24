from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .rpa_input_status import RPA_INPUT_SHEETS, TRANSACTION_UID_COLUMN, find_header_columns


RPA_TASKS_SHEET_NAME = "RPA_TASKS"


@dataclass
class SanitizedInputSummary:
    input_path: Path
    removed_rows: list[dict[str, Any]] = field(default_factory=list)
    removed_tasks: list[dict[str, Any]] = field(default_factory=list)
    updated_tasks: int = 0
    saved: bool = False

    @property
    def has_changes(self) -> bool:
        return bool(self.removed_rows or self.removed_tasks or self.updated_tasks)

    def lines(self) -> list[str]:
        return [
            f"Removed {len(self.removed_rows)} blank or invalid input rows.",
            f"Removed {len(self.removed_tasks)} stale RPA_TASKS rows.",
            f"Updated {self.updated_tasks} RPA_TASKS row mappings.",
        ]

    def format(self) -> str:
        return "\n".join(self.lines())


def sanitize_rpa_input(input_file: str | Path) -> SanitizedInputSummary:
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"RPA input file not found: {input_path}")

    workbook = load_workbook(input_path)
    summary = SanitizedInputSummary(input_path=input_path)

    for sheet_name in RPA_INPUT_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        headers = find_header_columns(ws)
        uid_column = headers.get(TRANSACTION_UID_COLUMN)
        if not uid_column:
            continue

        for row_index in range(ws.max_row, 1, -1):
            reason = _removal_reason(ws, row_index, uid_column)
            if not reason:
                continue
            summary.removed_rows.append({"sheet": ws.title, "row": row_index, "reason": reason})
            ws.delete_rows(row_index, 1)

    _sync_rpa_tasks(workbook, summary)

    if summary.has_changes:
        workbook.save(input_path)
        summary.saved = True

    return summary


def _removal_reason(ws, row_index: int, uid_column: int) -> str:
    row_values = [ws.cell(row=row_index, column=column_index).value for column_index in range(1, ws.max_column + 1)]
    if not any(_has_value(value) for value in row_values):
        return "blank"
    if not _clean_text(ws.cell(row=row_index, column=uid_column).value):
        return "missing_transaction_uid"
    return ""


def _sync_rpa_tasks(workbook, summary: SanitizedInputSummary) -> None:
    if RPA_TASKS_SHEET_NAME not in workbook.sheetnames:
        return

    input_rows, sheet_row_counts = _input_rows_by_uid(workbook)
    ws = workbook[RPA_TASKS_SHEET_NAME]
    headers = find_header_columns(ws)
    uid_column = headers.get(TRANSACTION_UID_COLUMN)
    if not uid_column:
        return

    for row_index in range(ws.max_row, 1, -1):
        uid = _clean_text(ws.cell(row=row_index, column=uid_column).value)
        if uid and uid in input_rows:
            continue
        summary.removed_tasks.append({"row": row_index, "transaction_uid": uid})
        ws.delete_rows(row_index, 1)

    updated_rows = 0
    for row_index in range(2, ws.max_row + 1):
        uid = _clean_text(ws.cell(row=row_index, column=uid_column).value)
        row_info = input_rows.get(uid)
        if not row_info:
            continue

        changed = False
        changed |= _set_task_cell(ws, headers, row_index, "input_sheet", row_info["input_sheet"])
        changed |= _set_task_cell(ws, headers, row_index, "input_excel_row", row_info["input_excel_row"])
        changed |= _set_task_cell(ws, headers, row_index, "sheet_row_count", sheet_row_counts[row_info["input_sheet"]])

        run_id_column = headers.get("run_id")
        task_id_column = headers.get("task_id")
        if run_id_column and task_id_column:
            run_id = _clean_text(ws.cell(row=row_index, column=run_id_column).value)
            if run_id:
                task_id = f"{run_id}:{row_info['input_sheet']}:{row_info['input_excel_row']}"
                changed |= _set_task_cell(ws, headers, row_index, "task_id", task_id)

        if changed:
            updated_rows += 1

    summary.updated_tasks = updated_rows


def _input_rows_by_uid(workbook) -> tuple[dict[str, dict[str, Any]], dict[str, int]]:
    rows: dict[str, dict[str, Any]] = {}
    sheet_row_counts: dict[str, int] = {}

    for sheet_name in RPA_INPUT_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        headers = find_header_columns(ws)
        uid_column = headers.get(TRANSACTION_UID_COLUMN)
        if not uid_column:
            sheet_row_counts[sheet_name] = 0
            continue

        count = 0
        for row_index in range(2, ws.max_row + 1):
            uid = _clean_text(ws.cell(row=row_index, column=uid_column).value)
            if not uid:
                continue
            rows[uid] = {"input_sheet": sheet_name, "input_excel_row": row_index}
            count += 1
        sheet_row_counts[sheet_name] = count

    return rows, sheet_row_counts


def _set_task_cell(ws, headers: dict[str, int], row_index: int, column_name: str, value: Any) -> bool:
    column_index = headers.get(column_name)
    if not column_index:
        return False
    cell = ws.cell(row=row_index, column=column_index)
    if cell.value == value:
        return False
    cell.value = value
    return True


def _has_value(value: Any) -> bool:
    return _clean_text(value) != ""


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
