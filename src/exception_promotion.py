from __future__ import annotations

import hashlib
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .config_loader import load_config
from .excel_io import atomic_output_path
from .flows import FLOW_BAO_CO, FLOW_BAO_NO, FLOW_CHI_TIEN_MAT, FLOW_THU_TIEN_MAT, flow_sheet
from .output_writer import (
    EXCEPTION_REVIEW_COLUMNS,
    RPA_REASON_ENCODING_TCVN3,
    RPA_REASON_UNICODE_COLUMN,
)
from .rpa_input_status import (
    INPUT_MESSAGE_COLUMN,
    INPUT_STATUS_COLUMN,
    INPUT_UPDATED_AT_COLUMN,
    RPA_INPUT_SHEETS,
    ensure_input_status_columns,
)
from .rpa_summary import SUMMARY_COLUMNS, load_summary, write_summary
from .rpa_tracking import STATUS_PENDING
from .vietnamese_encoding import unicode_to_tcvn3


EXCEPTION_SHEET_NAME = "EXCEPTION"
APPROVAL_COLUMN = "Duyệt nhập RPA"
FLOW_COLUMN = "Luồng nhập RPA"
EXCEPTION_STATUS_COLUMN = "Trạng thái xử lý"
MISSING_ERROR_COLUMN = "Vấn đề cần xử lý"
LEGACY_EXCEPTION_STATUS_COLUMN = "Trạng thái xử lý exception"
LEGACY_MISSING_ERROR_COLUMN = "Lỗi còn thiếu"
PROMOTED_TO_INPUT_COLUMN = "Promoted to input"
PROMOTED_SHEET_COLUMN = "Promoted sheet"
PROMOTED_AT_COLUMN = "Promoted at"

STATUS_PROMOTED = "Đã chuyển"
STATUS_MISSING_DATA = "Thiếu dữ liệu"
STATUS_INVALID_FLOW = "Luồng không hợp lệ"
LEGACY_STATUS_PROMOTED = "da_chuyen_sang_input"

FLOW_TO_SHEET = {
    FLOW_BAO_NO: flow_sheet(FLOW_BAO_NO),
    FLOW_BAO_CO: flow_sheet(FLOW_BAO_CO),
    FLOW_CHI_TIEN_MAT: flow_sheet(FLOW_CHI_TIEN_MAT),
    FLOW_THU_TIEN_MAT: flow_sheet(FLOW_THU_TIEN_MAT),
}

BASE_REQUIRED_COLUMNS = [
    "Ngày CT",
    "Mã ĐT",
    "TK nợ",
    "TK có",
    "Thành tiền",
]
FLOW_REQUIRED_COLUMNS = {
    FLOW_BAO_NO: BASE_REQUIRED_COLUMNS,
    FLOW_BAO_CO: BASE_REQUIRED_COLUMNS,
    FLOW_THU_TIEN_MAT: ["Ngày CT", "Người nộp/nhận tiền", "Mã ĐT", "TK nợ", "TK có", "Thành tiền"],
    FLOW_CHI_TIEN_MAT: ["Ngày CT", "Người nộp/nhận tiền", "Mã ĐT", "TK nợ", "TK có", "Thành tiền"],
}
TARGET_TRACKING_COLUMNS = [
    "transaction_uid",
    "run_id",
    INPUT_STATUS_COLUMN,
    INPUT_MESSAGE_COLUMN,
    INPUT_UPDATED_AT_COLUMN,
]
MISSING_SENTINELS = {"", "ERROR", "N/A", "NA", "NONE", "NULL"}


class ExceptionPromotionError(RuntimeError):
    """Raised when the workbook cannot be promoted safely."""


@dataclass
class PromotedRecord:
    transaction_uid: str
    flow: str
    input_sheet: str
    input_excel_row: int
    run_id: str
    exception_row: int
    row_data: dict[str, Any]
    promoted_at: str


@dataclass
class PromotionSummary:
    promoted: int = 0
    skipped_already_promoted: int = 0
    failed_validation: int = 0
    ignored: int = 0
    promoted_records: list[PromotedRecord] = field(default_factory=list)

    def lines(self) -> list[str]:
        return [
            f"Promoted {self.promoted} exception rows to input sheets.",
            f"Skipped {self.skipped_already_promoted} already promoted rows.",
            f"Failed validation {self.failed_validation} rows.",
        ]

    def format(self) -> str:
        return "\n".join(self.lines())


def promote_reviewed_exceptions(
    input_file: str | Path,
    *,
    config_path: str | Path | None = None,
) -> PromotionSummary:
    input_path = Path(input_file)
    workbook = load_workbook_for_promotion(input_path)
    if EXCEPTION_SHEET_NAME not in workbook.sheetnames:
        raise ExceptionPromotionError(f"Workbook does not contain sheet {EXCEPTION_SHEET_NAME}")

    normalize_legacy_cash_headers(workbook)
    exception_ws = workbook[EXCEPTION_SHEET_NAME]
    exception_headers = ensure_exception_columns(exception_ws)
    reason_encoding = load_rpa_reason_encoding(config_path)
    workbook_run_id = infer_workbook_run_id(workbook)
    existing_uids = collect_existing_uids(workbook, include_exception=False)
    known_uids = collect_existing_uids(workbook, include_exception=True)
    summary = PromotionSummary()

    for row_index in range(2, exception_ws.max_row + 1):
        row_data = normalize_exception_row(read_row(exception_ws, exception_headers, row_index))
        if not normalize_yes(row_data.get(APPROVAL_COLUMN)):
            summary.ignored += 1
            continue

        if is_already_promoted(row_data):
            summary.skipped_already_promoted += 1
            continue

        flow = normalize_flow(row_data.get(FLOW_COLUMN))
        target_sheet = map_flow_to_target_sheet(flow)
        if not target_sheet:
            write_validation_error(
                exception_ws,
                exception_headers,
                row_index,
                STATUS_INVALID_FLOW,
                [invalid_flow_message(row_data.get(FLOW_COLUMN))],
            )
            summary.failed_validation += 1
            continue

        if target_sheet not in workbook.sheetnames:
            raise ExceptionPromotionError(f"Workbook does not contain target sheet {target_sheet}")

        validation_errors = validate_exception_row(row_data, flow)
        if validation_errors:
            write_validation_error(
                exception_ws,
                exception_headers,
                row_index,
                STATUS_MISSING_DATA,
                validation_errors,
            )
            summary.failed_validation += 1
            continue

        target_ws = workbook[target_sheet]
        target_headers = ensure_target_tracking_columns(target_ws)
        transaction_uid = ensure_transaction_uid(row_data, row_index, known_uids)
        run_id = clean_text(row_data.get("run_id")) or workbook_run_id
        promoted_at = now()
        if transaction_uid in existing_uids:
            write_promotion_success(
                exception_ws,
                exception_headers,
                row_index,
                transaction_uid=transaction_uid,
                run_id=run_id,
                target_sheet=target_sheet,
                promoted_at=promoted_at,
            )
            summary.skipped_already_promoted += 1
            continue
        input_excel_row = promote_row(
            source_row=row_data,
            target_ws=target_ws,
            target_headers=target_headers,
            transaction_uid=transaction_uid,
            run_id=run_id,
            reason_encoding=reason_encoding,
            flow=flow,
        )
        write_promotion_success(
            exception_ws,
            exception_headers,
            row_index,
            transaction_uid=transaction_uid,
            run_id=run_id,
            target_sheet=target_sheet,
            promoted_at=promoted_at,
        )
        summary.promoted += 1
        summary.promoted_records.append(
            PromotedRecord(
                transaction_uid=transaction_uid,
                flow=flow,
                input_sheet=target_sheet,
                input_excel_row=input_excel_row,
                run_id=run_id,
                exception_row=row_index,
                row_data={**row_data, "transaction_uid": transaction_uid, "run_id": run_id},
                promoted_at=promoted_at,
            )
        )
        existing_uids.add(transaction_uid)
        known_uids.add(transaction_uid)

    with atomic_output_path(input_path) as temporary_path:
        workbook.save(temporary_path)
    ensure_summary_rows(input_path, summary.promoted_records)
    return summary


def load_workbook_for_promotion(input_file: Path):
    if not input_file.exists():
        raise FileNotFoundError(f"RPA input file not found: {input_file}")
    return load_workbook(input_file)


def ensure_exception_columns(ws) -> dict[str, int]:
    headers = find_header_columns(ws)
    is_new_schema = EXCEPTION_STATUS_COLUMN in headers or "Vấn đề cần xử lý" in headers
    if is_new_schema:
        required_columns = [
            "transaction_uid",
            "run_id",
            "source_file",
            "source_sheet",
            "source_row",
            *EXCEPTION_REVIEW_COLUMNS,
        ]
    else:
        required_columns = [
            "transaction_uid",
            "run_id",
            APPROVAL_COLUMN,
            FLOW_COLUMN,
            LEGACY_EXCEPTION_STATUS_COLUMN,
            LEGACY_MISSING_ERROR_COLUMN,
            PROMOTED_TO_INPUT_COLUMN,
            PROMOTED_SHEET_COLUMN,
            PROMOTED_AT_COLUMN,
        ]
    return ensure_columns(ws, required_columns)


def ensure_target_tracking_columns(ws) -> dict[str, int]:
    headers = ensure_columns(ws, ["transaction_uid", "run_id"])
    headers = ensure_input_status_columns(ws)
    for column_name in TARGET_TRACKING_COLUMNS:
        if column_name not in headers:
            headers = ensure_columns(ws, [column_name])
    return headers


def ensure_columns(ws, required_columns: list[str]) -> dict[str, int]:
    headers = find_header_columns(ws)
    last_column = max(headers.values()) if headers else 0
    for column_name in required_columns:
        if column_name in headers:
            continue
        last_column += 1
        ws.cell(row=1, column=last_column, value=column_name)
        copy_header_style(ws, last_column - 1, last_column)
        headers[column_name] = last_column
    return headers


def find_header_columns(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for index, cell in enumerate(ws[1], start=1):
        name = clean_text(cell.value)
        if name and name not in headers:
            headers[name] = index
    return headers


def read_row(ws, headers: dict[str, int], row_index: int) -> dict[str, Any]:
    return {
        column_name: clean_cell(ws.cell(row=row_index, column=column_index).value)
        for column_name, column_index in headers.items()
    }


def normalize_legacy_cash_headers(workbook) -> None:
    mappings = {
        flow_sheet(FLOW_THU_TIEN_MAT): ("Người nhận tiền", "Người nộp tiền"),
        flow_sheet(FLOW_CHI_TIEN_MAT): ("Người nộp tiền", "Người nhận tiền"),
    }
    for sheet_name, (legacy_name, correct_name) in mappings.items():
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        headers = find_header_columns(ws)
        if correct_name not in headers and legacy_name in headers:
            ws.cell(row=1, column=headers[legacy_name], value=correct_name)


def normalize_exception_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    normalized["transaction_uid"] = clean_text(row.get("transaction_uid")) or clean_text(row.get("Mã định danh"))
    normalized["source_file"] = clean_text(row.get("source_file")) or clean_text(row.get("File gốc"))
    normalized["source_sheet"] = clean_text(row.get("source_sheet")) or clean_text(row.get("Sheet gốc"))
    normalized["source_row"] = row.get("source_row") or row.get("Dòng gốc") or ""
    normalized["Nội dung giao dịch"] = clean_text(row.get("Nội dung giao dịch")) or clean_text(
        row.get("Nội dung giao dịch gốc")
    )
    normalized["Đối tượng giao dịch"] = clean_text(row.get("Đối tượng giao dịch")) or clean_text(
        row.get("Người hưởng/Người chuyển")
    )
    flow = normalize_flow(row.get(FLOW_COLUMN) or row.get("Luồng"))
    legacy_person = (
        clean_text(row.get("Người nhận tiền"))
        if flow == FLOW_THU_TIEN_MAT
        else clean_text(row.get("Người nộp tiền"))
        if flow == FLOW_CHI_TIEN_MAT
        else ""
    )
    normalized["Người nộp/nhận tiền"] = (
        clean_text(row.get("Người nộp/nhận tiền"))
        or legacy_person
        or clean_text(row.get("Người nộp tiền"))
        or clean_text(row.get("Người nhận tiền"))
    )
    normalized["Tên ĐT"] = clean_text(row.get("Tên ĐT")) or clean_text(row.get("Tên ĐT suy luận"))
    normalized["Lý do"] = (
        clean_text(row.get("Lý do"))
        or clean_text(row.get(RPA_REASON_UNICODE_COLUMN))
        or clean_text(row.get("Lí do"))
    )
    return normalized


def normalize_yes(value: Any) -> bool:
    return clean_text(value).lower() == "yes"


def normalize_flow(value: Any) -> str:
    return clean_text(value).lower().replace("-", "_").replace(" ", "_")


def map_flow_to_target_sheet(flow: str) -> str:
    return FLOW_TO_SHEET.get(flow, "")


def is_already_promoted(row_data: dict[str, Any]) -> bool:
    status = clean_text(row_data.get(EXCEPTION_STATUS_COLUMN) or row_data.get(LEGACY_EXCEPTION_STATUS_COLUMN))
    return (
        normalize_yes(row_data.get(PROMOTED_TO_INPUT_COLUMN))
        or status.lower() == LEGACY_STATUS_PROMOTED
        or status.startswith(STATUS_PROMOTED)
    )


def validate_exception_row(row_data: dict[str, Any], flow: str) -> list[str]:
    errors = [f"Thiếu {column_name}" for column_name in FLOW_REQUIRED_COLUMNS[flow] if is_missing(row_data.get(column_name))]
    if not has_reason(row_data):
        errors.append(f"Thiếu Lí do hoặc {RPA_REASON_UNICODE_COLUMN}")
    if is_foreign_currency_row(row_data) and is_missing(row_data.get("Tỷ giá")):
        errors.append("Thiếu Tỷ giá")
    return errors


def has_reason(row_data: dict[str, Any]) -> bool:
    return not is_missing(row_data.get("Lý do"))


def is_foreign_currency_row(row_data: dict[str, Any]) -> bool:
    for column_name in ("Ngoại tệ", "Số tiền ngoại tệ", "foreign_currency", "foreign_amount"):
        if not is_missing(row_data.get(column_name)):
            return True
    for column_name in ("TK nợ", "TK có"):
        if clean_text(row_data.get(column_name)).startswith("1122"):
            return True
    return False


def is_missing(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip().upper() in MISSING_SENTINELS
    return False


def promote_row(
    *,
    source_row: dict[str, Any],
    target_ws,
    target_headers: dict[str, int],
    transaction_uid: str,
    run_id: str,
    reason_encoding: str,
    flow: str,
) -> int:
    target_row_index = target_ws.max_row + 1
    copy_row_style(target_ws, target_row_index - 1, target_row_index)

    reason_unicode = clean_text(source_row.get("Lý do"))
    reason_for_rpa = encode_reason(reason_unicode, reason_encoding)

    for column_name, column_index in target_headers.items():
        value = source_row.get(column_name, "")
        if column_name == "transaction_uid":
            value = transaction_uid
        elif column_name == "run_id":
            value = run_id
        elif column_name == INPUT_STATUS_COLUMN:
            value = STATUS_PENDING
        elif column_name == INPUT_MESSAGE_COLUMN:
            value = ""
        elif column_name == INPUT_UPDATED_AT_COLUMN:
            value = ""
        elif column_name == "Lí do":
            value = reason_for_rpa
        elif column_name == RPA_REASON_UNICODE_COLUMN:
            value = reason_unicode
        elif column_name == "Người nộp tiền" and flow == FLOW_THU_TIEN_MAT:
            value = source_row.get("Người nộp/nhận tiền", "")
        elif column_name == "Người nhận tiền" and flow == FLOW_CHI_TIEN_MAT:
            value = source_row.get("Người nộp/nhận tiền", "")
        target_ws.cell(row=target_row_index, column=column_index, value=value)

    apply_number_formats(target_ws, target_headers, target_row_index)
    return target_row_index


def write_validation_error(
    ws,
    headers: dict[str, int],
    row_index: int,
    status: str,
    errors: list[str],
) -> None:
    status_column = EXCEPTION_STATUS_COLUMN if EXCEPTION_STATUS_COLUMN in headers else LEGACY_EXCEPTION_STATUS_COLUMN
    issue_column = MISSING_ERROR_COLUMN if MISSING_ERROR_COLUMN in headers else LEGACY_MISSING_ERROR_COLUMN
    legacy_status = {
        STATUS_MISSING_DATA: "thieu_du_lieu",
        STATUS_INVALID_FLOW: "luong_khong_hop_le",
    }.get(status, status)
    ws.cell(row=row_index, column=headers[status_column], value=status if status_column == EXCEPTION_STATUS_COLUMN else legacy_status)
    ws.cell(row=row_index, column=headers[issue_column], value="; ".join(errors))


def write_promotion_success(
    ws,
    headers: dict[str, int],
    row_index: int,
    *,
    transaction_uid: str,
    run_id: str,
    target_sheet: str,
    promoted_at: str,
) -> None:
    if EXCEPTION_STATUS_COLUMN in headers:
        display_time = promoted_at.replace("T", " ")
        ws.cell(
            row=row_index,
            column=headers[EXCEPTION_STATUS_COLUMN],
            value=f"{STATUS_PROMOTED} vào {target_sheet} lúc {display_time}",
        )
        ws.cell(row=row_index, column=headers[MISSING_ERROR_COLUMN], value="")
    else:
        ws.cell(row=row_index, column=headers[LEGACY_EXCEPTION_STATUS_COLUMN], value=LEGACY_STATUS_PROMOTED)
        ws.cell(row=row_index, column=headers[LEGACY_MISSING_ERROR_COLUMN], value="")
        ws.cell(row=row_index, column=headers[PROMOTED_TO_INPUT_COLUMN], value="yes")
        ws.cell(row=row_index, column=headers[PROMOTED_SHEET_COLUMN], value=target_sheet)
        ws.cell(row=row_index, column=headers[PROMOTED_AT_COLUMN], value=promoted_at)
    if "transaction_uid" in headers:
        ws.cell(row=row_index, column=headers["transaction_uid"], value=transaction_uid)
    if "Mã định danh" in headers:
        ws.cell(row=row_index, column=headers["Mã định danh"], value=transaction_uid)
    if "run_id" in headers:
        ws.cell(row=row_index, column=headers["run_id"], value=run_id)


def ensure_transaction_uid(row_data: dict[str, Any], row_index: int, existing_uids: set[str]) -> str:
    candidate = clean_text(row_data.get("transaction_uid"))
    if candidate:
        return candidate

    fingerprint_parts = [
        str(row_index),
        clean_text(row_data.get("source_file")),
        clean_text(row_data.get("source_sheet")),
        clean_text(row_data.get("source_row")),
        clean_text(row_data.get("Ngày CT")),
        clean_text(row_data.get("Thành tiền")),
        clean_text(row_data.get("Nội dung giao dịch")),
        clean_text(row_data.get(FLOW_COLUMN)),
    ]
    digest = hashlib.sha1("|".join(fingerprint_parts).encode("utf-8")).hexdigest()[:16]
    base_uid = f"exception_{digest}"
    uid = base_uid
    suffix = 2
    while uid in existing_uids:
        uid = f"{base_uid}_{suffix}"
        suffix += 1
    return uid


def collect_existing_uids(workbook, *, include_exception: bool = True) -> set[str]:
    uids: set[str] = set()
    sheet_names = list(RPA_INPUT_SHEETS)
    if include_exception:
        sheet_names.append(EXCEPTION_SHEET_NAME)
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        headers = find_header_columns(ws)
        uid_columns = [headers[column] for column in ("transaction_uid", "Mã định danh") if column in headers]
        for row_index in range(2, ws.max_row + 1):
            for column_index in uid_columns:
                uid = clean_text(ws.cell(row=row_index, column=column_index).value)
                if uid:
                    uids.add(uid)
    return uids


def infer_workbook_run_id(workbook) -> str:
    for sheet_name in [*RPA_INPUT_SHEETS, "RPA_TASKS", EXCEPTION_SHEET_NAME]:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        headers = find_header_columns(ws)
        run_id_column = headers.get("run_id")
        if not run_id_column:
            continue
        for row_index in range(2, ws.max_row + 1):
            run_id = clean_text(ws.cell(row=row_index, column=run_id_column).value)
            if run_id:
                return run_id
    return ""


def ensure_summary_rows(input_path: Path, promoted_records: list[PromotedRecord]) -> None:
    if not promoted_records:
        return
    summary_path = input_path.parent / "rpa_summary.xlsx"
    if not summary_path.exists():
        return

    summary_df = load_summary(summary_path)
    existing_uids = {
        clean_text(value)
        for value in summary_df.get("transaction_uid", pd.Series(dtype=object)).tolist()
        if clean_text(value)
    }
    new_rows = [summary_record_from_promoted(record) for record in promoted_records if record.transaction_uid not in existing_uids]
    if not new_rows:
        return

    updated_df = pd.concat([summary_df, pd.DataFrame(new_rows, columns=SUMMARY_COLUMNS)], ignore_index=True)
    write_summary(updated_df, summary_path)


def summary_record_from_promoted(record: PromotedRecord) -> dict[str, Any]:
    row = {column: "" for column in SUMMARY_COLUMNS}
    data = record.row_data
    source_row = clean_text(data.get("source_row")) or record.exception_row
    bank = clean_text(data.get("Ngân hàng"))
    source_file = clean_text(data.get("source_file"))
    source_sheet = clean_text(data.get("source_sheet")) or EXCEPTION_SHEET_NAME
    row.update(
        {
            "Ngày CT": data.get("Ngày CT") or "",
            "Ngân hàng": bank,
            "Luồng": record.flow,
            "Nội dung giao dịch": clean_text(data.get("Nội dung giao dịch")),
            "Mã ĐT": clean_text(data.get("Mã ĐT")),
            "Tên ĐT": clean_text(data.get("Tên ĐT")),
            "TK nợ": clean_text(data.get("TK nợ")),
            "TK có": clean_text(data.get("TK có")),
            "Thành tiền": data.get("Thành tiền") or "",
            "Kết quả phân loại": STATUS_PROMOTED,
            "Trạng thái RPA": STATUS_PENDING,
            "Số chứng từ VACOM": "",
            "Thông báo RPA": "",
            "Thời gian hoàn thành": "",
            "Nguồn sao kê": f"{source_file} | {source_sheet} | dòng {source_row}",
            "transaction_uid": record.transaction_uid,
            "last_run_id": record.run_id,
            "last_attempt_result": "",
            "rpa_started_at": "",
            "rpa_finished_at": "",
        }
    )
    return row


def load_rpa_reason_encoding(config_path: str | Path | None = None) -> str:
    path = Path(config_path) if config_path else Path(__file__).resolve().parents[1] / "config" / "config.yaml"
    try:
        config = load_config(path)
    except OSError:
        return ""
    return clean_text(config.get("output", {}).get("rpa_reason_encoding"))


def encode_reason(reason: str, reason_encoding: str) -> str:
    if clean_text(reason_encoding).lower() == RPA_REASON_ENCODING_TCVN3:
        return unicode_to_tcvn3(reason)
    return reason


def invalid_flow_message(flow_value: Any) -> str:
    value = clean_text(flow_value)
    if not value:
        return f"Thiếu {FLOW_COLUMN}"
    return f"{FLOW_COLUMN} không hợp lệ: {value}"


def copy_header_style(ws, source_column: int, target_column: int) -> None:
    if source_column < 1:
        return
    source = ws.cell(row=1, column=source_column)
    target = ws.cell(row=1, column=target_column)
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    if source_row < 2:
        return
    for column_index in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=column_index)
        target = ws.cell(row=target_row, column=column_index)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
            target.number_format = source.number_format


def apply_number_formats(ws, headers: dict[str, int], row_index: int) -> None:
    for column_name, column_index in headers.items():
        cell = ws.cell(row=row_index, column=column_index)
        if "Ngày" in column_name:
            cell.number_format = "DD/MM/YYYY"
        elif column_name in {"Thành tiền", "Tỷ giá"}:
            cell.number_format = "#,##0"


def clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def clean_text(value: Any) -> str:
    value = clean_cell(value)
    if value is None:
        return ""
    return str(value).strip()


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")
